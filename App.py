###############################################################
# RXTRACK: EXECUTIVE DASHBOARD (v13.4 - Stability Fixes)
# Architecture: Quad-Table Strategy + Attendance + Pricing
# Updates:
#   1. Fixed Sidebar Indentation & Duplicate Logic.
#   2. Implemented Day/Week/Month/Range Date Filters.
###############################################################

import streamlit as st
import pandas as pd
import numpy as np
import hashlib
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta, date
import gc
import re
import io
import warnings
import json
import time
import hmac
from openpyxl import load_workbook

from sqlalchemy import text
import os

from rxtrack_shared import (
    _DEFAULT_ADMIN_USERS,
    db_cursor,
    engine,
    execute_statement,
    generate_pk,
    load_admin_users,
    normalize_identifier_text,
    normalize_name,
    parse_shift_start,
    seconds_to_mmss,
)


def init_db():
    """Initializes tables if they do not exist."""
    schemas = [
        """CREATE TABLE IF NOT EXISTS events (
            pk TEXT PRIMARY KEY, user_name TEXT, device TEXT, med_id TEXT, med_desc TEXT, 
            event_type TEXT, dt TIMESTAMP, qty FLOAT, beginning_qty FLOAT, ending_qty FLOAT, 
            discrepancy_qty FLOAT, discrepancy_reason TEXT, resolution_dt TIMESTAMP
        );""",
        """CREATE TABLE IF NOT EXISTS config_events (
            pk TEXT PRIMARY KEY, dt TIMESTAMP, user_name TEXT, device TEXT, med_id TEXT, 
            location TEXT, action_type TEXT, activity_category TEXT, min_qty FLOAT, max_qty FLOAT, is_standard BOOLEAN
        );""",
        """CREATE TABLE IF NOT EXISTS med_costs (
            med_id TEXT PRIMARY KEY, cost_per_unit FLOAT
        );""",
        """CREATE TABLE IF NOT EXISTS pharmacy_orders (
            pk TEXT PRIMARY KEY, queue_id TEXT, priority TEXT, dt TIMESTAMP, med_id TEXT, 
            med_desc TEXT, destination TEXT, user_name TEXT, qty FLOAT
        );""",
        """CREATE TABLE IF NOT EXISTS staff_schedule (
            pk TEXT PRIMARY KEY, dt DATE, day_name TEXT, staff_name TEXT, 
            shift_type TEXT, assignment_type TEXT, raw_entry TEXT, note TEXT,
            schedule_status TEXT, cell_fill_color TEXT
        );""",
        """CREATE TABLE IF NOT EXISTS attendance_punches (
            pk TEXT PRIMARY KEY, raw_name TEXT, dt_date DATE, start_dt TIMESTAMP, end_dt TIMESTAMP
        );""",
        """CREATE TABLE IF NOT EXISTS inventory_audit (
            pk TEXT PRIMARY KEY, med_id TEXT, med_desc TEXT, med_class TEXT, 
            unit_cost FLOAT, qty_on_hand FLOAT, min_lvl FLOAT, max_lvl FLOAT
        );""",
        """CREATE TABLE IF NOT EXISTS inventory_detailed (
            pk TEXT PRIMARY KEY, station TEXT, med_id TEXT, med_desc TEXT,
            unit_cost FLOAT, current_count FLOAT, pocket_location TEXT
        );""",
        """CREATE TABLE IF NOT EXISTS cycle_count_status (
            pk TEXT PRIMARY KEY,
            snapshot_date DATE,
            source_filename TEXT,
            isa_name TEXT,
            med_id TEXT,
            med_desc TEXT,
            location TEXT,
            cycle_count_interval FLOAT,
            last_cycle_count TIMESTAMP,
            days_since_last_count FLOAT,
            days_over_due FLOAT
        );""",
        """CREATE TABLE IF NOT EXISTS cycle_count_variances (
            pk TEXT PRIMARY KEY,
            variance_type TEXT,
            dt TIMESTAMP,
            med_id TEXT,
            med_desc TEXT,
            starting_qty FLOAT,
            new_qty FLOAT,
            qty_variance FLOAT,
            unit_cost FLOAT,
            extended_cost FLOAT,
            user_name TEXT,
            source_filename TEXT
        );""",
        """CREATE TABLE IF NOT EXISTS buyer_formulary_listing (
            pk TEXT PRIMARY KEY,
            snapshot_date DATE,
            source_filename TEXT,
            location TEXT,
            med_id TEXT,
            med_desc TEXT,
            package_size FLOAT,
            min_qty FLOAT,
            max_qty FLOAT,
            qty FLOAT,
            ndc TEXT,
            distributor TEXT,
            item_code TEXT,
            purchase_date DATE,
            received_qty FLOAT,
            uploaded_at TIMESTAMP DEFAULT NOW()
        );""",
        """CREATE TABLE IF NOT EXISTS physical_inventory_snapshots (
            pk TEXT PRIMARY KEY,
            snapshot_date DATE,
            source_filename TEXT,
            isa_name TEXT,
            med_id TEXT,
            med_desc TEXT,
            min_qty FLOAT,
            max_qty FLOAT,
            on_hand_qty FLOAT,
            location TEXT,
            unit_cost FLOAT,
            extended_cost FLOAT,
            uploaded_at TIMESTAMP DEFAULT NOW()
        );""",
        """CREATE TABLE IF NOT EXISTS audit_transaction_detail_rc (
            pk TEXT PRIMARY KEY,
            care_area_name TEXT,
            location TEXT,
            station_name TEXT,
            source_system TEXT,
            dt TIMESTAMP,
            user_name TEXT,
            user_id TEXT,
            user_type TEXT,
            priority_code TEXT,
            transaction_type TEXT,
            med_id TEXT,
            med_desc TEXT,
            generic_name TEXT,
            med_class TEXT,
            therapeutic_class TEXT,
            drawer_subdrawer_pocket TEXT,
            min_qty FLOAT,
            max_qty FLOAT,
            dispense_amount FLOAT,
            qty FLOAT,
            beginning_qty FLOAT,
            ending_qty FLOAT,
            unit_cost FLOAT,
            extended_cost FLOAT,
            discrepancy TEXT,
            discrepancy_difference FLOAT,
            discrepancy_resolution_desc TEXT,
            discrepancy_reason TEXT,
            correction_quantity_before FLOAT,
            correction_quantity_after FLOAT,
            correction TEXT,
            resolution_user TEXT,
            resolution_dt TIMESTAMP,
            waste_amount FLOAT,
            waste_reason TEXT,
            witness_user_name TEXT,
            override_reason TEXT,
            override_flag TEXT,
            ordering_physician_present BOOLEAN,
            attending_physician_present BOOLEAN,
            source_filename TEXT,
            uploaded_at TIMESTAMP DEFAULT NOW()
        );""",
        """CREATE TABLE IF NOT EXISTS packaged_meds (
            pk TEXT PRIMARY KEY,
            dispense_dt TIMESTAMP,
            reception_num TEXT,
            med_id TEXT,
            med_desc TEXT,
            dose_form TEXT,
            qty_per_pack FLOAT,
            qoh FLOAT,
            manufacturer TEXT,
            ndc TEXT,
            mfg_lot_number TEXT,
            mfg_expire_date DATE,
            device_id TEXT,
            hospital_lot_number TEXT,
            hospital_expire_date DATE,
            bud DATE,
            packaged_by TEXT,
            confirmer TEXT
        );""",
        """CREATE TABLE IF NOT EXISTS device_inventory (
            pk TEXT PRIMARY KEY,
            med_desc TEXT,
            device TEXT,
            zone TEXT,
            pocket_location TEXT,
            status TEXT,
            brand_name TEXT,
            med_id TEXT,
            med_class TEXT,
            current_quantity FLOAT,
            min_qty FLOAT,
            max_qty FLOAT,
            outdate_tracking TEXT,
            loaded_as_fraction TEXT,
            backordered TEXT,
            standard_stock TEXT,
            active_orders TEXT,
            days_unused FLOAT,
            snapshot_dt TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );""",
        """CREATE TABLE IF NOT EXISTS device_inventory_history (
            snapshot_date DATE NOT NULL DEFAULT CURRENT_DATE,
            pk TEXT NOT NULL,
            med_desc TEXT,
            device TEXT,
            zone TEXT,
            pocket_location TEXT,
            status TEXT,
            brand_name TEXT,
            med_id TEXT,
            med_class TEXT,
            current_quantity FLOAT,
            min_qty FLOAT,
            max_qty FLOAT,
            outdate_tracking TEXT,
            loaded_as_fraction TEXT,
            backordered TEXT,
            standard_stock TEXT,
            active_orders TEXT,
            days_unused FLOAT,
            snapshot_dt TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );""",
        """CREATE UNIQUE INDEX IF NOT EXISTS idx_device_inventory_history_snapshot_pk
            ON device_inventory_history (snapshot_date, pk);""",
        """CREATE TABLE IF NOT EXISTS iv_room_workload (
            pk TEXT PRIMARY KEY,
            facility_name TEXT,
            order_lot_number TEXT,
            compound_type TEXT,
            num_preparations FLOAT,
            dose_number TEXT,
            drug_name TEXT,
            order_date DATE,
            ordered_time TEXT,
            order_dt TIMESTAMP,
            completed_on TIMESTAMP,
            priority_name TEXT,
            prepare_tat_minutes FLOAT,
            prepared_by TEXT,
            approved_by TEXT,
            secondary_approved_by TEXT
        );""",
        """CREATE TABLE IF NOT EXISTS iv_room_workflow_detail (
            pk TEXT PRIMARY KEY,
            facility_name TEXT,
            order_lot_number TEXT,
            dose_number TEXT,
            ordered_on DATE,
            prepared_by TEXT,
            approved_by TEXT,
            drug_name TEXT,
            workflow_name TEXT,
            workflow_step_type TEXT,
            workflow_step_name TEXT,
            workflow_step_category TEXT,
            start_date DATE,
            start_time TEXT,
            stop_time TEXT,
            start_dt TIMESTAMP,
            stop_dt TIMESTAMP,
            total_duration_minutes FLOAT,
            source_file TEXT,
            uploaded_at TIMESTAMP DEFAULT NOW()
        );""",
        """CREATE INDEX IF NOT EXISTS idx_iv_room_workflow_detail_start_date
            ON iv_room_workflow_detail (start_date);""",
        """CREATE INDEX IF NOT EXISTS idx_iv_room_workflow_detail_order_lot
            ON iv_room_workflow_detail (order_lot_number, dose_number);""",
        """CREATE TABLE IF NOT EXISTS wcc_compounding_stats (
            pk TEXT PRIMARY KEY,
            component_name TEXT,
            component_id TEXT,
            order_name TEXT,
            administration_dt TIMESTAMP,
            barcode_status TEXT,
            source_file TEXT,
            uploaded_at TIMESTAMP DEFAULT NOW()
        );""",
        """CREATE TABLE IF NOT EXISTS wcc_cartfill_stats (
            pk TEXT PRIMARY KEY,
            report_start_date DATE,
            report_end_date DATE,
            order_medication TEXT,
            med_id TEXT,
            ready_for_dispense_dt TIMESTAMP,
            admin_given_dt TIMESTAMP,
            prepared_dt TIMESTAMP,
            prep_or_dispense_user TEXT,
            location TEXT,
            pharmacy TEXT,
            cartfill_area TEXT,
            source_file TEXT,
            uploaded_at TIMESTAMP DEFAULT NOW()
        );""",
        """CREATE TABLE IF NOT EXISTS overnight_iv_cartfill_orders (
            pk TEXT PRIMARY KEY,
            order_id TEXT,
            order_medication TEXT,
            ready_for_dispense_dt TIMESTAMP,
            admin_given_dt TIMESTAMP,
            prepared_dt TIMESTAMP,
            prep_or_dispense_user TEXT,
            pharmacy TEXT,
            event_date DATE,
            required_start_dt TIMESTAMP,
            prep_lead_hours FLOAT,
            hold_hours FLOAT,
            is_sjs_cleanroom BOOLEAN
        );""",
        """CREATE TABLE IF NOT EXISTS overnight_iv_cartfill_windows (
            pk TEXT PRIMARY KEY,
            cartfill_name TEXT,
            time_processed_raw TEXT,
            doses_due TEXT,
            pharmacy TEXT
        );""",
        """CREATE TABLE IF NOT EXISTS overnight_iv_staffing_model (
            pk TEXT PRIMARY KEY,
            schedule_date DATE,
            day_name TEXT,
            shift_name TEXT,
            weekend_shift_label TEXT,
            assigned_staff TEXT,
            is_weekend BOOLEAN,
            is_placeholder BOOLEAN
        );""",
        """CREATE TABLE IF NOT EXISTS admin_users (
            username TEXT PRIMARY KEY,
            display_name TEXT,
            added_at TIMESTAMP DEFAULT NOW()
        );""",
        """CREATE TABLE IF NOT EXISTS daily_ops (
            id SERIAL PRIMARY KEY,
            task TEXT NOT NULL,
            category TEXT,
            priority TEXT DEFAULT 'Medium',
            status TEXT DEFAULT 'Not Started',
            due_date DATE,
            created_at TIMESTAMP DEFAULT NOW(),
            notes TEXT
        );""",
        """CREATE TABLE IF NOT EXISTS follow_ups (
            id SERIAL PRIMARY KEY,
            item TEXT NOT NULL,
            action_taken TEXT,
            follow_up_date DATE,
            status TEXT DEFAULT 'Pending',
            notes TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );""",
        """CREATE TABLE IF NOT EXISTS recurring_tasks (
            id SERIAL PRIMARY KEY,
            task TEXT NOT NULL,
            category TEXT,
            priority TEXT DEFAULT 'Medium',
            recurrence TEXT DEFAULT 'Daily',
            days_of_week TEXT,
            active BOOLEAN DEFAULT TRUE,
            notes TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );""",
        """CREATE TABLE IF NOT EXISTS management_coaching_notes (
            id SERIAL PRIMARY KEY,
            staff_name TEXT NOT NULL,
            topic TEXT,
            coaching_date DATE,
            follow_up_date DATE,
            status TEXT DEFAULT 'Open',
            summary TEXT,
            next_steps TEXT,
            source_page TEXT,
            source_key TEXT,
            source_payload_json TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        );""",
        """CREATE TABLE IF NOT EXISTS shift_audit_profiles (
            profile_name TEXT PRIMARY KEY,
            page_name TEXT NOT NULL,
            shifts_json TEXT,
            selected_names_json TEXT,
            view_scope TEXT,
            active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        );""",
        """CREATE TABLE IF NOT EXISTS shift_audit_results (
            profile_name TEXT NOT NULL,
            audit_date DATE NOT NULL,
            shifts_json TEXT,
            selected_names_json TEXT,
            view_scope TEXT,
            staff_on_shift INTEGER,
            sessions INTEGER,
            active_sec FLOAT,
            walk_sec FLOAT,
            long_gap_count INTEGER,
            training_count INTEGER,
            dominant_work_type TEXT,
            work_type_json TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW(),
            PRIMARY KEY (profile_name, audit_date)
        );""",
        """ALTER TABLE daily_ops ADD COLUMN IF NOT EXISTS recurring_task_id INTEGER;""",
        """ALTER TABLE staff_schedule ADD COLUMN IF NOT EXISTS schedule_status TEXT;""",
        """ALTER TABLE staff_schedule ADD COLUMN IF NOT EXISTS cell_fill_color TEXT;""",
        """ALTER TABLE wcc_cartfill_stats ADD COLUMN IF NOT EXISTS cartfill_area TEXT;""",
        """ALTER TABLE wcc_cartfill_stats ADD COLUMN IF NOT EXISTS admin_given_dt TIMESTAMP;""",
        """ALTER TABLE cycle_count_variances ADD COLUMN IF NOT EXISTS source_filename TEXT;""",
        """ALTER TABLE management_coaching_notes ADD COLUMN IF NOT EXISTS source_page TEXT;""",
        """ALTER TABLE management_coaching_notes ADD COLUMN IF NOT EXISTS source_key TEXT;""",
        """ALTER TABLE management_coaching_notes ADD COLUMN IF NOT EXISTS source_payload_json TEXT;""",
        """CREATE UNIQUE INDEX IF NOT EXISTS idx_management_coaching_notes_source_key
            ON management_coaching_notes (source_key)
            WHERE source_key IS NOT NULL;"""
    ]
    with db_cursor() as (conn, cur):
        for sql in schemas:
            cur.execute(sql)
        conn.commit()

def run_query(query, params=None):
    """Executes a SELECT query and returns a pandas DataFrame."""
    try:
        with db_cursor() as (conn, cur):
            return pd.read_sql(query, conn, params=params)
    except Exception:
        return pd.DataFrame()

SHIFT_WORK_TYPE_ORDER = [
    "Carousel / 0400 Pull",
    "Pyxis Outdates",
    "Returns / Carousel Putaway",
    "Pyxis Maintenance",
]


@st.cache_data(ttl=300)
def load_shift_audit_profiles(page_name="shift_work_map"):
    try:
        init_db()
        sql = text("""
            SELECT profile_name, page_name, shifts_json, selected_names_json, view_scope, active
            FROM shift_audit_profiles
            WHERE page_name = :page_name AND active = TRUE
            ORDER BY profile_name
        """)
        with engine.connect() as conn:
            df = pd.read_sql(sql, conn, params={"page_name": page_name})
        if df.empty:
            return pd.DataFrame(columns=["profile_name", "page_name", "shifts", "selected_names", "view_scope", "active"])
        df["shifts"] = df["shifts_json"].apply(lambda x: json.loads(x) if pd.notna(x) and str(x).strip() else [])
        df["selected_names"] = df["selected_names_json"].apply(lambda x: json.loads(x) if pd.notna(x) and str(x).strip() else [])
        return df
    except Exception:
        return pd.DataFrame(columns=["profile_name", "page_name", "shifts", "selected_names", "view_scope", "active"])


@st.cache_data(ttl=300)
def load_shift_schedule_for_date(sel_date):
    try:
        sql = text("""
            SELECT pk, dt, day_name, staff_name, shift_type, assignment_type, raw_entry, note,
                   COALESCE(schedule_status, assignment_type, 'Standard') AS schedule_status,
                   cell_fill_color
            FROM staff_schedule
            WHERE dt = :d
        """)
        with engine.connect() as conn:
            df = pd.read_sql(sql, conn, params={"d": str(sel_date)})
        if df.empty:
            return df
        df["dt"] = pd.to_datetime(df["dt"]).dt.date
        df["staff_name"] = df["staff_name"].fillna("Unknown").astype(str).str.strip()
        df["shift_type"] = df["shift_type"].fillna("").astype(str).str.strip()
        df["assignment_type"] = df["assignment_type"].fillna("").astype(str).str.strip()
        df["schedule_status"] = df["schedule_status"].fillna("Standard").astype(str).str.strip()
        df["match_key"] = df["staff_name"].apply(normalize_name)
        return df
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=300)
def load_day_events_for_shift_audit(sel_date):
    try:
        sql = text("""
            SELECT pk, dt, user_name, device, event_type, med_desc, qty
            FROM events
            WHERE dt::date = :d
        """)
        with engine.connect() as conn:
            df = pd.read_sql(sql, conn, params={"d": str(sel_date)})
        if df.empty:
            return df
        df["dt"] = pd.to_datetime(df["dt"], errors="coerce")
        df["user_name"] = df["user_name"].fillna("Unknown").astype(str).str.strip()
        df["device"] = df["device"].fillna("Unknown").astype(str).str.strip()
        df["event_type"] = df["event_type"].fillna("").astype(str).str.strip()
        df["match_key"] = df["user_name"].apply(normalize_name)
        return df
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=300)
def load_day_pharmacy_for_shift_audit(sel_date):
    try:
        sql = text("""
            SELECT pk, dt, user_name, destination, priority, med_desc, qty
            FROM pharmacy_orders
            WHERE dt::date = :d
        """)
        with engine.connect() as conn:
            df = pd.read_sql(sql, conn, params={"d": str(sel_date)})
        if df.empty:
            return df
        df["dt"] = pd.to_datetime(df["dt"], errors="coerce")
        df["user_name"] = df["user_name"].fillna("Unknown").astype(str).str.strip()
        df["destination"] = df["destination"].fillna("Unknown").astype(str).str.strip()
        df["priority"] = df["priority"].fillna("").astype(str).str.strip()
        return df
    except Exception:
        return pd.DataFrame()


def classify_shift_work_type(source, device, event_type):
    src = str(source or "").lower()
    dev = str(device or "").lower()
    evt = str(event_type or "").lower()

    if src == "pharmacy":
        return "Carousel / 0400 Pull"
    if re.search(r"outdate", evt):
        return "Pyxis Outdates"
    if re.search(r"carousel|cubic|pack|central", dev):
        return "Returns / Carousel Putaway"
    return "Pyxis Maintenance"


def build_shift_day_sessions(df_events_day, df_pharm_day):
    px_df = df_events_day[["pk", "dt", "user_name", "device", "event_type", "med_desc", "qty", "match_key"]].copy() if not df_events_day.empty else pd.DataFrame()
    if not px_df.empty:
        px_df["source"] = "Pyxis"

    ph_df = df_pharm_day[["pk", "dt", "user_name", "destination", "priority", "med_desc", "qty"]].copy() if not df_pharm_day.empty else pd.DataFrame()
    if not ph_df.empty:
        ph_df = ph_df.rename(columns={"destination": "device", "priority": "event_type"})
        ph_df["source"] = "Pharmacy"
        ph_df["match_key"] = ph_df["user_name"].apply(normalize_name)

    day_combined = pd.concat([px_df, ph_df], ignore_index=True)
    if day_combined.empty:
        return pd.DataFrame(), pd.DataFrame()

    day_combined["dt"] = pd.to_datetime(day_combined["dt"], errors="coerce")
    day_combined = day_combined.dropna(subset=["dt"]).sort_values(["match_key", "dt"]).reset_index(drop=True)
    day_combined["prev_match_key"] = day_combined["match_key"].shift()
    day_combined["prev_device"] = day_combined["device"].shift()
    day_combined["prev_dt"] = day_combined["dt"].shift()
    day_combined["gap"] = (day_combined["dt"] - day_combined["prev_dt"]).dt.total_seconds().fillna(0)
    day_combined["is_new_session"] = np.where(
        (day_combined["match_key"] != day_combined["prev_match_key"]) |
        (day_combined["device"] != day_combined["prev_device"]) |
        (day_combined["gap"] > 1200),
        1, 0
    )
    day_combined["session_id"] = day_combined["is_new_session"].cumsum()

    grouped = (
        day_combined.groupby("session_id")
        .agg(
            tech_key=("match_key", "first"),
            user_name=("user_name", "first"),
            device=("device", "first"),
            source=("source", "first"),
            primary_event=("event_type", "first"),
            start=("dt", "min"),
            end=("dt", "max"),
            tx_count=("pk", "count"),
        )
        .reset_index()
    )
    grouped["duration_sec"] = (grouped["end"] - grouped["start"]).dt.total_seconds()
    grouped["duration_sec"] = np.where(grouped["duration_sec"] < 10, 30, grouped["duration_sec"])
    grouped = grouped.sort_values(["tech_key", "start"]).reset_index(drop=True)
    grouped["next_start"] = grouped.groupby("tech_key")["start"].shift(-1)
    grouped["walk_sec"] = (grouped["next_start"] - grouped["end"]).dt.total_seconds()
    grouped["walk_sec"] = grouped["walk_sec"].where(grouped["walk_sec"].gt(0), 0).fillna(0)
    return day_combined, grouped


def summarize_shift_audit_sessions(active_sessions, active_work_keys, training_count):
    return {
        "staff_on_shift": len(active_work_keys),
        "sessions": int(len(active_sessions)),
        "active_sec": float(active_sessions["duration_sec"].sum()),
        "walk_sec": float(active_sessions["walk_sec"].sum()),
        "long_gap_count": int(active_sessions["long_gap_flag"].sum()),
        "training_count": int(training_count),
    }


def run_shift_audit_profile_for_date(sel_date, shifts, selected_names=None, view_scope="Whole Shift Team"):
    selected_names = selected_names or []
    df_sched = load_shift_schedule_for_date(sel_date)
    if df_sched.empty:
        return None

    profile_shifts = [s for s in shifts if s in df_sched["shift_type"].dropna().unique()]
    if not profile_shifts:
        return None

    df_events = load_day_events_for_shift_audit(sel_date)
    df_pharm = load_day_pharmacy_for_shift_audit(sel_date)
    if df_events.empty and df_pharm.empty:
        return None

    profile_sched = df_sched[df_sched["shift_type"].isin(profile_shifts)].copy()
    if view_scope == "Selected Staff Only" and selected_names:
        profile_keys = set(profile_sched[profile_sched["staff_name"].isin(selected_names)]["match_key"].unique())
    else:
        profile_keys = set(profile_sched["match_key"].unique())
    if not profile_keys:
        return None

    day_stream, day_sessions = build_shift_day_sessions(df_events, df_pharm)
    if day_sessions.empty:
        return None

    active_sessions = day_sessions[day_sessions["tech_key"].isin(profile_keys)].copy()
    if active_sessions.empty:
        return None

    work_key_to_name = (
        profile_sched[["match_key", "staff_name", "assignment_type"]]
        .drop_duplicates("match_key")
        .set_index("match_key")
    )
    key_to_name = work_key_to_name["staff_name"].to_dict()
    key_to_assignment = work_key_to_name["assignment_type"].fillna("").astype(str).to_dict()

    active_sessions["assignment_type"] = active_sessions["tech_key"].map(key_to_assignment).fillna("")
    active_sessions["tech_display"] = active_sessions["tech_key"].map(key_to_name).fillna(active_sessions["user_name"])
    active_sessions["tech_display"] = np.where(
        active_sessions["assignment_type"].str.lower().eq("training"),
        active_sessions["tech_display"] + " (Training)",
        active_sessions["tech_display"],
    )
    active_sessions["work_type"] = active_sessions.apply(
        lambda row: classify_shift_work_type(row["source"], row["device"], row["primary_event"]),
        axis=1
    )
    active_sessions["work_type"] = pd.Categorical(
        active_sessions["work_type"],
        categories=SHIFT_WORK_TYPE_ORDER,
        ordered=True,
    )
    active_sessions["long_gap_flag"] = active_sessions["walk_sec"] > 1200

    training_count = int(
        profile_sched["assignment_type"].fillna("").astype(str).str.lower().eq("training").sum()
    )
    summary = summarize_shift_audit_sessions(active_sessions, profile_keys, training_count)

    work_type_breakdown = (
        active_sessions.groupby("work_type", observed=False, as_index=False)
        .agg(
            sessions=("session_id", "count"),
            active_sec=("duration_sec", "sum"),
            walk_sec=("walk_sec", "sum"),
            techs=("tech_display", "nunique"),
        )
        .sort_values("work_type")
    )
    work_type_breakdown = work_type_breakdown[work_type_breakdown["sessions"] > 0]
    dominant_work_type = (
        work_type_breakdown.sort_values(["active_sec", "work_type"], ascending=[False, True])["work_type"].iloc[0]
        if not work_type_breakdown.empty else ""
    )

    return {
        "audit_date": pd.to_datetime(sel_date).date(),
        "shifts": profile_shifts,
        "selected_names": selected_names,
        "view_scope": view_scope,
        "staff_on_shift": summary["staff_on_shift"],
        "sessions": summary["sessions"],
        "active_sec": summary["active_sec"],
        "walk_sec": summary["walk_sec"],
        "long_gap_count": summary["long_gap_count"],
        "training_count": summary["training_count"],
        "dominant_work_type": dominant_work_type,
        "work_type_breakdown": work_type_breakdown.to_dict("records"),
    }


def save_shift_audit_result(profile_name, audit_result):
    init_db()
    sql = text("""
        INSERT INTO shift_audit_results (
            profile_name, audit_date, shifts_json, selected_names_json, view_scope,
            staff_on_shift, sessions, active_sec, walk_sec, long_gap_count,
            training_count, dominant_work_type, work_type_json, updated_at
        )
        VALUES (
            :profile_name, :audit_date, :shifts_json, :selected_names_json, :view_scope,
            :staff_on_shift, :sessions, :active_sec, :walk_sec, :long_gap_count,
            :training_count, :dominant_work_type, :work_type_json, NOW()
        )
        ON CONFLICT (profile_name, audit_date) DO UPDATE SET
            shifts_json = EXCLUDED.shifts_json,
            selected_names_json = EXCLUDED.selected_names_json,
            view_scope = EXCLUDED.view_scope,
            staff_on_shift = EXCLUDED.staff_on_shift,
            sessions = EXCLUDED.sessions,
            active_sec = EXCLUDED.active_sec,
            walk_sec = EXCLUDED.walk_sec,
            long_gap_count = EXCLUDED.long_gap_count,
            training_count = EXCLUDED.training_count,
            dominant_work_type = EXCLUDED.dominant_work_type,
            work_type_json = EXCLUDED.work_type_json,
            updated_at = NOW()
    """)
    payload = {
        "profile_name": profile_name,
        "audit_date": audit_result["audit_date"],
        "shifts_json": json.dumps(audit_result["shifts"]),
        "selected_names_json": json.dumps(audit_result["selected_names"]),
        "view_scope": audit_result["view_scope"],
        "staff_on_shift": audit_result["staff_on_shift"],
        "sessions": audit_result["sessions"],
        "active_sec": audit_result["active_sec"],
        "walk_sec": audit_result["walk_sec"],
        "long_gap_count": audit_result["long_gap_count"],
        "training_count": audit_result["training_count"],
        "dominant_work_type": audit_result["dominant_work_type"],
        "work_type_json": json.dumps(audit_result["work_type_breakdown"]),
    }
    with engine.begin() as conn:
        conn.execute(sql, payload)


@st.cache_data(ttl=300)
def load_shift_audit_results(start_date, end_date, profile_names=None):
    init_db()
    if profile_names:
        sql = text("""
            SELECT *
            FROM shift_audit_results
            WHERE audit_date BETWEEN :start_date AND :end_date
              AND profile_name = ANY(:profile_names)
            ORDER BY audit_date, profile_name
        """)
        params = {
            "start_date": str(start_date),
            "end_date": str(end_date),
            "profile_names": list(profile_names),
        }
    else:
        sql = text("""
            SELECT *
            FROM shift_audit_results
            WHERE audit_date BETWEEN :start_date AND :end_date
            ORDER BY audit_date, profile_name
        """)
        params = {"start_date": str(start_date), "end_date": str(end_date)}

    with engine.connect() as conn:
        df = pd.read_sql(sql, conn, params=params)
    if df.empty:
        return df
    df["audit_date"] = pd.to_datetime(df["audit_date"]).dt.date
    df["shifts"] = df["shifts_json"].apply(lambda x: json.loads(x) if pd.notna(x) and str(x).strip() else [])
    df["selected_names"] = df["selected_names_json"].apply(lambda x: json.loads(x) if pd.notna(x) and str(x).strip() else [])
    df["work_type_breakdown"] = df["work_type_json"].apply(lambda x: json.loads(x) if pd.notna(x) and str(x).strip() else [])
    return df


def excel_serial_to_datetime(series):
    fallback = pd.to_datetime(series, errors="coerce")
    numeric = pd.to_numeric(series, errors="coerce")
    plausible_excel_serial = numeric.between(1, 100000)
    converted = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
    if plausible_excel_serial.any():
        converted.loc[plausible_excel_serial] = (
            pd.Timestamp("1899-12-30")
            + pd.to_timedelta(numeric.loc[plausible_excel_serial], unit="D")
        )
    return converted.where(plausible_excel_serial, fallback)

# --- DATA CLEANING ---
def clean_dataframe(df):
    df = df.copy()
    colmap = {
        "UserName": "user_name", "UserID": "user_id", "Device": "device",
        "MedID": "med_id", "MedDescription": "med_desc", "TransactionType": "event_type",
        "TransactionDateTime": "dt", "Quantity": "qty", "Beg": "beginning_qty", 
        "End": "ending_qty", "DiscrepancyQuantity": "discrepancy_qty", 
        "DiscrepancyReason": "discrepancy_reason", "ResolutionDatetime": "resolution_dt"
    }
    df.rename(columns=colmap, inplace=True)
    required = ["user_name", "device", "med_id", "med_desc", "event_type", "dt", "qty", 
                "beginning_qty", "ending_qty", "discrepancy_qty", "discrepancy_reason", "resolution_dt"]
    for col in required:
        if col not in df.columns: df[col] = None
    df["dt"] = pd.to_datetime(df["dt"], errors="coerce")
    df.dropna(subset=["dt"], inplace=True)
    df["resolution_dt"] = pd.to_datetime(df["resolution_dt"], errors="coerce")
    for c in ["qty", "discrepancy_qty", "beginning_qty", "ending_qty"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype('float32')
    df["med_id"] = df["med_id"].fillna("").astype(str).str.strip().str.upper().replace({"NAN": ""})
    # Leave timestamps as Python datetimes for psycopg2 and convert missing values to SQL NULL.
    df["discrepancy_reason"] = df["discrepancy_reason"].where(pd.notna(df["discrepancy_reason"]), None)
    df["resolution_dt"] = df["resolution_dt"].where(pd.notna(df["resolution_dt"]), None)
    df = df.astype(object)
    df = df.where(pd.notna(df), None)
    df["pk"] = df.apply(generate_pk, axis=1)
    return df[required + ["pk"]]

def clean_activity_log(df):
    df = df.copy()
    df.columns = df.columns.str.strip().str.replace(' ', '')
    df.rename(columns={
        "UserName": "user_name", "Device": "device", "TransactionDateTime": "dt", 
        "Action": "action_type", "ActivityType": "activity_category", "AffectedElement": "raw_element",
        "Amount": "qty_col", "Quantity": "qty_col"
    }, inplace=True)
    df["dt"] = pd.to_datetime(df["dt"], errors="coerce")
    df.dropna(subset=["dt"], inplace=True)
    pattern_element = r'^(.*?) \((.*?)\)'
    extracted = df['raw_element'].astype(str).str.extract(pattern_element)
    df['location'] = extracted[0].str.strip()
    df['med_id'] = extracted[1].str.strip()
    df.dropna(subset=['med_id'], inplace=True)
    if 'qty_col' not in df.columns:
        pattern_qty = r':\s*(\d+)$' 
        df['qty_col'] = df['raw_element'].astype(str).str.extract(pattern_qty)[0]
    df['qty_extracted'] = pd.to_numeric(df['qty_col'], errors='coerce')
    df.sort_values(['user_name', 'device', 'med_id', 'dt'], inplace=True)
    df['time_gap'] = df.groupby(['user_name', 'device', 'med_id'])['dt'].diff().dt.total_seconds().fillna(999)
    df['group_id'] = (df['time_gap'] > 120).astype(int).cumsum()
    df['is_min'] = df['activity_category'].str.contains('Min', case=False, na=False)
    df['is_max'] = df['activity_category'].str.contains('Max', case=False, na=False)
    df['is_std'] = df['activity_category'].str.contains('Standard Stock', case=False, na=False)
    df['min_qty'] = np.where(df['is_min'], df['qty_extracted'], np.nan)
    df['max_qty'] = np.where(df['is_max'], df['qty_extracted'], np.nan)
    df['max_qty'] = np.where((~df['is_min']) & (~df['is_max']), df['qty_extracted'], df['max_qty'])
    grouped = df.groupby(['user_name', 'device', 'med_id', 'group_id'], as_index=False).agg({
        'min_qty': 'max', 'max_qty': 'max', 'is_std': 'max',
        'location': 'first', 'dt': 'first', 'action_type': 'first', 'activity_category': 'first'
    })
    grouped["dt"] = grouped["dt"].astype(str)
    grouped["pk"] = grouped.apply(generate_pk, axis=1)
    grouped.replace({np.nan: None}, inplace=True)
    return grouped.rename(columns={'is_std': 'is_standard'})[['pk', 'dt', 'user_name', 'device', 'med_id', 'location', 'action_type', 'activity_category', 'min_qty', 'max_qty', 'is_standard']]

def clean_pharmacy_report(df):
    df = df.copy()
    colmap = {
        "TranQueueID": "queue_id", "Priority": "priority", "Date / Time": "dt",
        "Item ID": "med_id", "Description": "med_desc", "Destination": "destination",
        "User": "user_name", "Quantity": "qty"
    }
    df.rename(columns=colmap, inplace=True)
    for col in colmap.values():
        if col not in df.columns: df[col] = None
    df["dt"] = pd.to_datetime(df["dt"], errors="coerce")
    df.dropna(subset=["dt"], inplace=True)
    df["qty"] = pd.to_numeric(df["qty"], errors="coerce").fillna(0)
    df["dt"] = df["dt"].astype(str)
    df["pk"] = df.apply(generate_pk, axis=1)
    return df[["pk", "queue_id", "priority", "dt", "med_id", "med_desc", "destination", "user_name", "qty"]]

def normalize_fill_color(fill):
    """Return ARGB/RGB fill value from an openpyxl PatternFill, when present."""
    if fill is None or getattr(fill, "fill_type", None) is None:
        return ""
    color = fill.fgColor
    if not color:
        return ""
    if color.type == "rgb" and color.rgb:
        return str(color.rgb).upper()
    if color.type == "indexed":
        return f"INDEXED:{color.indexed}"
    if color.type == "theme":
        return f"THEME:{color.theme}"
    return ""


def classify_schedule_status(raw_entry, fill_color=""):
    text = str(raw_entry or "").lower()
    if "trade" in text:
        return "Trade"
    if "incentive" in text or "bonus" in text:
        return "Incentive Pay"
    if "out early" in text or " in @" in text or "out @" in text:
        return "Adjustment"
    if "adjust" in text:
        return "Adjustment"
    if "open" in text or text.strip() in {"op", "open shift"}:
        return "Open Shift"

    rgb = str(fill_color or "").replace("#", "").upper()
    if len(rgb) == 8:
        rgb = rgb[-6:]
    if len(rgb) != 6 or not re.fullmatch(r"[0-9A-F]{6}", rgb):
        return "Standard"

    r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)

    # Full-row weekend colors in the schedule should not become exceptions.
    if g > r and g > b and g >= 150:
        return "Standard"
    if r >= 230 and g >= 170 and b >= 180:
        return "Standard"

    # Gold/yellow/orange cells in the schedule are used for shift adjustments.
    if r >= 220 and g >= 140 and b <= 90:
        return "Adjustment"
    if r > 210 and 80 <= g <= 190 and b < 120:
        return "Adjustment"
    if r >= 170 and g < 105 and b < 105:
        return "Open Shift"
    if 80 <= r < 170 and g < 80 and b < 80:
        return "Incentive Pay"
    return "Standard"


def schedule_fill_lookup_from_workbook(file_obj):
    try:
        file_obj.seek(0)
        wb = load_workbook(file_obj, data_only=True)
        ws = wb.active
        lookup = {}
        for row in ws.iter_rows():
            date_val = row[1].value if len(row) > 1 else None
            schedule_date = excel_serial_to_datetime(pd.Series([date_val])).iloc[0]
            if pd.isna(schedule_date):
                continue
            for cell in row[3:]:
                if cell.value is None:
                    continue
                lookup[(schedule_date.date(), str(cell.value).strip())] = normalize_fill_color(cell.fill)
        return lookup
    except Exception:
        return {}


def clean_schedule_data(df, schedule_file=None):
    df = df.copy()
    fill_lookup = schedule_fill_lookup_from_workbook(schedule_file) if schedule_file is not None else {}
    if len(df.columns) > 2:
        df.rename(columns={df.columns[1]: 'Date', df.columns[2]: 'Day'}, inplace=True)
    df = df.iloc[1:].dropna(subset=['Date'])
    df.drop(columns=[df.columns[0]], errors='ignore', inplace=True)
    long_df = df.melt(id_vars=['Date', 'Day'], var_name='col_header', value_name='raw_entry')
    long_df.dropna(subset=['raw_entry'], inplace=True)
    long_df = long_df[~long_df['raw_entry'].astype(str).str.lower().isin(['x', 'nan', '', ' '])]
    
    processed_rows = []
    for _, row in long_df.iterrows():
        raw = str(row['raw_entry']).strip()
        header = str(row['col_header']).strip()
        dt = pd.to_datetime(row['Date'], errors='coerce').date()
        day_name = row['Day']
        
        if re.search(r'\(\d', raw): 
            parts = [p.strip() + ')' for p in raw.split(')') if '(' in p]
        else:
            parts = [p.strip() for p in raw.split('\n') if p.strip()]
        
        for part in parts:
            if not part or part == ')': continue
            fill_color = fill_lookup.get((dt, part), fill_lookup.get((dt, raw), ""))
            schedule_status = classify_schedule_status(part, fill_color)
            override_time = None
            m_range = re.search(r'\(?(\d{4})\s*-\s*\d{4}\)?', part)
            if m_range:
                override_time = m_range.group(1)
                clean_part = part.replace(m_range.group(0), '')
            else:
                m_single = re.search(r'\((\d{4})\)', part)
                if m_single:
                    override_time = m_single.group(1)
                    clean_part = part.replace(m_single.group(0), '')
                else:
                    m_short = re.search(r'\((\d{1,2})\s*-\s*\d{1,2}\)', part)
                    if m_short:
                        override_time = m_short.group(1)
                        clean_part = part.replace(m_short.group(0), '')
                    else:
                        clean_part = part
            clean_part = clean_part.replace('()', '').strip()
            if clean_part.endswith(','): clean_part = clean_part[:-1]
            assignment_type = "Shift"
            note = ""
            lower_part = clean_part.lower()
            if 'trn' in lower_part or 'training' in lower_part:
                assignment_type = "Training"
                clean_part = re.split(r'\s(?:trn|training)\s?', clean_part, flags=re.IGNORECASE)[0].strip()
            elif any(x in lower_part for x in ['pto', 'off', 'sick']):
                assignment_type = "PTO"
            elif schedule_status != "Standard":
                assignment_type = schedule_status
            clean_part = re.sub(r'\s*\((trade|adjustment|adjust|incentive|bonus|open shift|open)\)\s*', ' ', clean_part, flags=re.IGNORECASE).strip()
            final_shift_str = override_time if override_time else header
            row_str = f"{dt}|{clean_part}|{final_shift_str}|{schedule_status}"
            pk = hashlib.sha256(row_str.encode()).hexdigest()
            processed_rows.append({
                'pk': pk, 'dt': dt, 'day_name': day_name, 'staff_name': clean_part.title(),
                'shift_type': final_shift_str, 'assignment_type': assignment_type,
                'raw_entry': part, 'note': note, 'schedule_status': schedule_status,
                'cell_fill_color': fill_color,
            })
    return pd.DataFrame(processed_rows)

def clean_attendance_file(file_obj):
    file_obj.seek(0)
    content = file_obj.read().decode('utf-8', errors='ignore')
    lines = content.splitlines()
    data = []
    name_pat = re.compile(r'Employee:\s*([A-Za-z\-,\s\.]+?)(?="|",|",Date)')
    date_pat = re.compile(r'Date:\s*(\d{1,2}/\d{1,2}/\d{4})')
    time_pat = re.compile(r'(\d{1,2}/\d{1,2}/\d{4}\s+\d{1,2}:\d{2})')
    for line in lines:
        if "Employee:" not in line or "Date:" not in line: continue
        m_name = name_pat.search(line)
        name = m_name.group(1).strip() if m_name else None
        m_date = date_pat.search(line)
        date_str = m_date.group(1) if m_date else None
        times = time_pat.findall(line)
        start_time = times[0] if len(times) > 0 else None
        end_time = times[1] if len(times) > 1 else None
        if name and date_str and start_time:
            data.append({"raw_name": name, "dt_date": pd.to_datetime(date_str).date(), "start_dt": start_time, "end_dt": end_time})
    df = pd.DataFrame(data)
    if not df.empty: df["pk"] = df.apply(generate_pk, axis=1)
    return df

def clean_inventory_file(df):
    df = df.copy()
    colmap = {"MedID": "med_id", "MedDescription": "med_desc", "MedClass": "med_class", "UnitCost": "unit_cost", "CurrentCount": "qty_on_hand", "CurrentMin": "min_lvl", "CurrentMax": "max_lvl"}
    df.rename(columns=colmap, inplace=True)
    for c in ["med_id", "med_desc", "unit_cost", "qty_on_hand", "min_lvl", "max_lvl"]:
        if c not in df.columns: df[c] = None
    if df['unit_cost'].dtype == object:
        df['unit_cost'] = df['unit_cost'].astype(str).str.replace('$', '', regex=False).str.replace(',', '', regex=False)
    df['unit_cost'] = pd.to_numeric(df['unit_cost'], errors='coerce').fillna(0)
    df['qty_on_hand'] = pd.to_numeric(df['qty_on_hand'], errors='coerce').fillna(0)
    df['med_id'] = df['med_id'].fillna("").astype(str).str.strip().str.upper().replace({"NAN": ""})
    df['pk'] = df.apply(lambda x: str(x['med_id']), axis=1)
    return df[['pk', 'med_id', 'med_desc', 'med_class', 'unit_cost', 'qty_on_hand', 'min_lvl', 'max_lvl']]

def clean_detailed_inventory(df):
    df = df.copy()
    colmap = {
        "StationName": "station", 
        "SourceSystem": "source_system",
        "MedID": "med_id", 
        "MedDescription": "med_desc", 
        "UnitCost": "unit_cost", 
        "CurrentCount": "current_count", 
        "DrawerSubdrawerPocket": "pocket_location"
    }
    df.rename(columns=colmap, inplace=True)
    required = ["station", "source_system", "med_id", "med_desc", "unit_cost", "current_count", "pocket_location"]
    for c in required:
        if c not in df.columns: df[c] = None
    if df['unit_cost'].dtype == object:
        df['unit_cost'] = df['unit_cost'].astype(str).str.replace('$', '', regex=False).str.replace(',', '', regex=False)
    df['unit_cost'] = pd.to_numeric(df['unit_cost'], errors='coerce').fillna(0)
    df['current_count'] = pd.to_numeric(df['current_count'], errors='coerce').fillna(0)
    df['med_id'] = df['med_id'].fillna("").astype(str).str.strip().str.upper().replace({"NAN": ""})
    df['row_sig'] = df['station'].astype(str) + df['med_id'].astype(str) + df['pocket_location'].astype(str)
    df['pk'] = df['row_sig'].apply(lambda x: hashlib.sha256(x.encode()).hexdigest())
    return df[required + ['pk']]


def clean_cycle_count_status_report(file_obj):
    file_obj.seek(0)
    metadata_line = file_obj.readline()
    if isinstance(metadata_line, bytes):
        metadata_line = metadata_line.decode("utf-8", errors="ignore")

    file_obj.seek(0)
    try:
        df = pd.read_csv(file_obj, header=1)
    except UnicodeDecodeError:
        file_obj.seek(0)
        df = pd.read_csv(file_obj, header=1, encoding="latin1")

    df = df.copy()
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(" ", "_")
        .str.replace("/", "_")
    )
    df = df.rename(columns={
        "isa_name": "isa_name",
        "med_id": "med_id",
        "description": "med_desc",
        "location": "location",
        "cycle_count_interval": "cycle_count_interval",
        "last_cycle_count": "last_cycle_count",
        "days_since_last_count": "days_since_last_count",
        "days_over_due": "days_over_due",
    })

    required = [
        "isa_name", "med_id", "med_desc", "location",
        "cycle_count_interval", "last_cycle_count",
        "days_since_last_count", "days_over_due",
    ]
    for col in required:
        if col not in df.columns:
            df[col] = None

    filename = getattr(file_obj, "name", "") or ""
    snapshot_date = None
    m_fname = re.search(r"(\d{8})", filename)
    if m_fname:
        snapshot_date = pd.to_datetime(m_fname.group(1), format="%m%d%Y", errors="coerce")
    if pd.isna(snapshot_date) or snapshot_date is None:
        m_meta = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", str(metadata_line))
        if m_meta:
            snapshot_date = pd.to_datetime(m_meta.group(1), errors="coerce")
    if pd.isna(snapshot_date) or snapshot_date is None:
        snapshot_date = pd.Timestamp.today().normalize()

    df["source_filename"] = filename
    df["snapshot_date"] = pd.to_datetime(snapshot_date, errors="coerce").date()
    df["isa_name"] = df["isa_name"].fillna("").astype(str).str.strip()
    df["med_id"] = df["med_id"].fillna("").astype(str).str.strip().str.upper()
    df["med_desc"] = df["med_desc"].fillna("").astype(str).str.strip()
    df["location"] = df["location"].fillna("").astype(str).str.strip().str.upper()
    df["cycle_count_interval"] = pd.to_numeric(df["cycle_count_interval"], errors="coerce").fillna(0)
    df["days_since_last_count"] = pd.to_numeric(df["days_since_last_count"], errors="coerce").fillna(0)
    df["days_over_due"] = pd.to_numeric(df["days_over_due"], errors="coerce").fillna(0)

    last_count = pd.to_datetime(df["last_cycle_count"], errors="coerce")
    sentinel_mask = last_count.dt.year.fillna(0).le(1)
    df["last_cycle_count"] = last_count.mask(sentinel_mask)

    df = df[(df["med_id"] != "") & (df["location"] != "")].copy()
    df["pk"] = df.apply(
        lambda row: hashlib.sha256(
            "|".join([
                str(row["snapshot_date"]),
                str(row["isa_name"]),
                str(row["med_id"]),
                str(row["location"]),
            ]).encode()
        ).hexdigest(),
        axis=1,
    )

    df = df.astype(object)
    df = df.where(pd.notna(df), None)
    return df[["pk", "snapshot_date", "source_filename"] + required]


def _normalize_report_columns(df):
    df = df.copy()
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(".", "", regex=False)
        .str.replace(" ", "_", regex=False)
        .str.replace("/", "_", regex=False)
    )
    return df


def _read_metadata_header_csv(file_obj, dtype=None):
    file_obj.seek(0)
    metadata_line = file_obj.readline()
    if isinstance(metadata_line, bytes):
        metadata_line = metadata_line.decode("utf-8", errors="ignore")

    file_obj.seek(0)
    try:
        df = pd.read_csv(file_obj, header=1, dtype=dtype, low_memory=False)
    except UnicodeDecodeError:
        file_obj.seek(0)
        df = pd.read_csv(file_obj, header=1, dtype=dtype, encoding="latin1", low_memory=False)
    return df, metadata_line


def _snapshot_date_from_report(file_obj, metadata_line):
    filename = getattr(file_obj, "name", "") or ""
    snapshot_date = None
    m_fname = re.search(r"(\d{8})", filename)
    if m_fname:
        snapshot_date = pd.to_datetime(m_fname.group(1), format="%m%d%Y", errors="coerce")
    if pd.isna(snapshot_date) or snapshot_date is None:
        m_meta = re.search(r"(?:Report On:|Reporting Between\s+)?(\d{1,2}/\d{1,2}/\d{4})", str(metadata_line))
        if m_meta:
            snapshot_date = pd.to_datetime(m_meta.group(1), errors="coerce")
    if pd.isna(snapshot_date) or snapshot_date is None:
        snapshot_date = pd.Timestamp.today().normalize()
    return pd.to_datetime(snapshot_date, errors="coerce").date()


def _money_to_numeric(series):
    return pd.to_numeric(
        series.astype(str).str.replace("$", "", regex=False).str.replace(",", "", regex=False),
        errors="coerce",
    ).fillna(0)


def clean_cycle_count_variance_report(file_obj):
    df, _metadata_line = _read_metadata_header_csv(file_obj)
    df = _normalize_report_columns(df)
    df = df.rename(columns={
        "grptype": "variance_type",
        "date_time": "dt",
        "item_id": "med_id",
        "description": "med_desc",
        "starting_qoh": "starting_qty",
        "new_qoh": "new_qty",
        "quantity_variance": "qty_variance",
        "cost": "unit_cost",
        "extended_cost": "extended_cost",
        "user": "user_name",
    })

    required = [
        "variance_type", "dt", "med_id", "med_desc", "starting_qty",
        "new_qty", "qty_variance", "unit_cost", "extended_cost", "user_name",
    ]
    for col in required:
        if col not in df.columns:
            df[col] = None

    df["dt"] = pd.to_datetime(df["dt"], errors="coerce")
    df["med_id"] = df["med_id"].fillna("").astype(str).str.strip().str.upper()
    df["med_desc"] = df["med_desc"].fillna("").astype(str).str.strip()
    df["variance_type"] = df["variance_type"].fillna("").astype(str).str.strip()
    df["user_name"] = df["user_name"].fillna("").astype(str).str.strip()
    for col in ["starting_qty", "new_qty", "qty_variance", "unit_cost", "extended_cost"]:
        df[col] = _money_to_numeric(df[col])

    df = df[df["dt"].notna() & df["med_id"].ne("")].copy()
    df["source_filename"] = getattr(file_obj, "name", "") or ""
    df["pk"] = df.apply(
        lambda row: hashlib.sha256(
            "|".join([
                str(row["dt"]),
                str(row["variance_type"]),
                str(row["med_id"]),
                str(row["qty_variance"]),
                str(row["user_name"]),
            ]).encode()
        ).hexdigest(),
        axis=1,
    )
    df = df.astype(object).where(pd.notna(df), None)
    return df[["pk"] + required + ["source_filename"]]


def clean_buyer_formulary_listing_report(file_obj):
    df, metadata_line = _read_metadata_header_csv(file_obj, dtype=str)
    df = _normalize_report_columns(df)
    df = df.rename(columns={
        "location": "location",
        "med_id": "med_id",
        "description": "med_desc",
        "package_size": "package_size",
        "min": "min_qty",
        "max": "max_qty",
        "qty": "qty",
        "ndc": "ndc",
        "distributor": "distributor",
        "item_code": "item_code",
        "purchase_date": "purchase_date",
        "rcvqty": "received_qty",
    })

    required = [
        "location", "med_id", "med_desc", "package_size", "min_qty",
        "max_qty", "qty", "ndc", "distributor", "item_code",
        "purchase_date", "received_qty",
    ]
    for col in required:
        if col not in df.columns:
            df[col] = None

    df["snapshot_date"] = _snapshot_date_from_report(file_obj, metadata_line)
    df["source_filename"] = getattr(file_obj, "name", "") or ""
    for col in ["location", "med_id", "med_desc", "ndc", "distributor", "item_code"]:
        df[col] = df[col].fillna("").astype(str).str.strip()
    df["location"] = df["location"].str.upper()
    df["med_id"] = df["med_id"].str.upper()
    for col in ["package_size", "min_qty", "max_qty", "qty", "received_qty"]:
        df[col] = _money_to_numeric(df[col])
    df["purchase_date"] = pd.to_datetime(df["purchase_date"], errors="coerce").dt.date

    df = df[df["location"].ne("") & df["med_id"].ne("")].copy()
    df["pk"] = df.apply(
        lambda row: hashlib.sha256(
            "|".join([
                str(row["snapshot_date"]),
                str(row["location"]),
                str(row["med_id"]),
                str(row["ndc"]),
                str(row["item_code"]),
                str(row["purchase_date"]),
            ]).encode()
        ).hexdigest(),
        axis=1,
    )
    df = df.astype(object).where(pd.notna(df), None)
    return df[["pk", "snapshot_date", "source_filename"] + required]


def clean_physical_inventory_report(file_obj):
    df, metadata_line = _read_metadata_header_csv(file_obj, dtype=str)
    df = _normalize_report_columns(df)
    df = df.rename(columns={
        "isa_name": "isa_name",
        "med_id": "med_id",
        "item_description": "med_desc",
        "min": "min_qty",
        "max": "max_qty",
        "on_hand": "on_hand_qty",
        "location": "location",
        "cost": "unit_cost",
        "ext_cost": "extended_cost",
    })

    required = [
        "isa_name", "med_id", "med_desc", "min_qty", "max_qty",
        "on_hand_qty", "location", "unit_cost", "extended_cost",
    ]
    for col in required:
        if col not in df.columns:
            df[col] = None

    df["snapshot_date"] = _snapshot_date_from_report(file_obj, metadata_line)
    df["source_filename"] = getattr(file_obj, "name", "") or ""
    for col in ["isa_name", "med_id", "med_desc", "location"]:
        df[col] = df[col].fillna("").astype(str).str.strip()
    df["med_id"] = df["med_id"].str.upper()
    df["location"] = df["location"].str.upper()
    for col in ["min_qty", "max_qty", "on_hand_qty", "unit_cost", "extended_cost"]:
        df[col] = _money_to_numeric(df[col])

    df = df[df["isa_name"].ne("") & df["med_id"].ne("") & df["location"].ne("")].copy()
    df["pk"] = df.apply(
        lambda row: hashlib.sha256(
            "|".join([
                str(row["snapshot_date"]),
                str(row["isa_name"]),
                str(row["med_id"]),
                str(row["location"]),
            ]).encode()
        ).hexdigest(),
        axis=1,
    )
    df = df.astype(object).where(pd.notna(df), None)
    return df[["pk", "snapshot_date", "source_filename"] + required]


def clean_audit_transaction_detail_rc(df, source_filename=""):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    colmap = {
        "CareAreaName": "care_area_name",
        "Location": "location",
        "StationName": "station_name",
        "SourceSystem": "source_system",
        "TransactionDateTime": "dt",
        "UserName": "user_name",
        "UserID": "user_id",
        "UserType": "user_type",
        "PriorityCode": "priority_code",
        "TransactionType": "transaction_type",
        "MedID": "med_id",
        "MedDescription": "med_desc",
        "GenericName": "generic_name",
        "MedClass": "med_class",
        "TherapeuticClass": "therapeutic_class",
        "DrawerSubDrawerPocket": "drawer_subdrawer_pocket",
        "Min": "min_qty",
        "Max": "max_qty",
        "DispenseAmount": "dispense_amount",
        "Quantity": "qty",
        "BeginCount": "beginning_qty",
        "EndCount": "ending_qty",
        "UnitCost": "unit_cost",
        "ExtendedCost": "extended_cost",
        "Discrepancy": "discrepancy",
        "DiscrepancyDifference": "discrepancy_difference",
        "DiscrepancyResolutionDesc": "discrepancy_resolution_desc",
        "DiscrepancyReason": "discrepancy_reason",
        "CorrectionQuantityBefore": "correction_quantity_before",
        "CorrectionQuantityAfter": "correction_quantity_after",
        "Correction": "correction",
        "ResolutionUser": "resolution_user",
        "ResolutionDateTime": "resolution_dt",
        "WasteAmount": "waste_amount",
        "WasteReason": "waste_reason",
        "WitnessUserName": "witness_user_name",
        "OverrideReason": "override_reason",
        "Override": "override_flag",
        "OrderingPhysician": "ordering_physician",
        "AttendingPhysician": "attending_physician",
    }
    df = df.rename(columns=colmap)
    required = [
        "care_area_name", "location", "station_name", "source_system", "dt",
        "user_name", "user_id", "user_type", "priority_code", "transaction_type",
        "med_id", "med_desc", "generic_name", "med_class", "therapeutic_class",
        "drawer_subdrawer_pocket", "min_qty", "max_qty", "dispense_amount", "qty",
        "beginning_qty", "ending_qty", "unit_cost", "extended_cost", "discrepancy",
        "discrepancy_difference", "discrepancy_resolution_desc", "discrepancy_reason",
        "correction_quantity_before", "correction_quantity_after", "correction",
        "resolution_user", "resolution_dt", "waste_amount", "waste_reason",
        "witness_user_name", "override_reason", "override_flag",
        "ordering_physician_present", "attending_physician_present", "source_filename",
    ]
    for col in required + ["ordering_physician", "attending_physician"]:
        if col not in df.columns:
            df[col] = None

    df["dt"] = pd.to_datetime(df["dt"], errors="coerce")
    df["resolution_dt"] = pd.to_datetime(df["resolution_dt"], errors="coerce")
    text_cols = [
        "care_area_name", "location", "station_name", "source_system", "user_name",
        "user_id", "user_type", "priority_code", "transaction_type", "med_id",
        "med_desc", "generic_name", "med_class", "therapeutic_class",
        "drawer_subdrawer_pocket", "discrepancy", "discrepancy_resolution_desc",
        "discrepancy_reason", "correction", "resolution_user", "waste_reason",
        "witness_user_name", "override_reason", "override_flag",
    ]
    for col in text_cols:
        df[col] = df[col].fillna("").astype(str).str.strip()
    df["med_id"] = df["med_id"].str.upper().replace({"NAN": ""})
    for col in [
        "min_qty", "max_qty", "dispense_amount", "qty", "beginning_qty", "ending_qty",
        "unit_cost", "extended_cost", "discrepancy_difference",
        "correction_quantity_before", "correction_quantity_after", "waste_amount",
    ]:
        df[col] = _money_to_numeric(df[col])
    df["ordering_physician_present"] = df["ordering_physician"].fillna("").astype(str).str.strip().ne("")
    df["attending_physician_present"] = df["attending_physician"].fillna("").astype(str).str.strip().ne("")
    df["source_filename"] = source_filename
    df = df[df["dt"].notna() & df["med_id"].ne("")].copy()
    df["pk"] = df.apply(
        lambda row: hashlib.sha256(
            "|".join([
                str(row["dt"]),
                str(row["station_name"]),
                str(row["med_id"]),
                str(row["transaction_type"]),
                str(row["user_id"] or row["user_name"]),
                str(row["qty"]),
                str(row["beginning_qty"]),
                str(row["ending_qty"]),
            ]).encode()
        ).hexdigest(),
        axis=1,
    )
    df = df.astype(object).where(pd.notna(df), None)
    return df[["pk"] + required]


def build_med_cost_updates_from_audit_detail(df):
    if df is None or df.empty or "med_id" not in df.columns or "unit_cost" not in df.columns:
        return pd.DataFrame(columns=["med_id", "cost_per_unit"])

    costs = df[["med_id", "unit_cost"] + (["dt"] if "dt" in df.columns else [])].copy()
    costs["med_id"] = costs["med_id"].fillna("").astype(str).str.strip().str.upper()
    costs["cost_per_unit"] = pd.to_numeric(costs["unit_cost"], errors="coerce")
    costs = costs[costs["med_id"].ne("") & costs["cost_per_unit"].gt(0)].copy()
    if costs.empty:
        return pd.DataFrame(columns=["med_id", "cost_per_unit"])

    if "dt" in costs.columns:
        costs["dt"] = pd.to_datetime(costs["dt"], errors="coerce")
        costs = costs.sort_values(["med_id", "dt"], na_position="first")
        costs = costs.drop_duplicates("med_id", keep="last")
    else:
        costs = costs.drop_duplicates("med_id", keep="last")

    costs = costs[["med_id", "cost_per_unit"]].astype(object).where(pd.notna(costs), None)
    return costs


def clean_packaging_report(file_obj):
    file_obj.seek(0)
    name = getattr(file_obj, "name", "packaging_report")

    if str(name).lower().endswith(".xlsx"):
        df = pd.read_excel(file_obj)
    else:
        raw_bytes = file_obj.read()
        text = raw_bytes.decode("utf-8-sig", errors="replace")
        lines = text.splitlines()
        header_idx = next(
            (idx for idx, line in enumerate(lines) if "Dispense Time" in line and "Med ID" in line),
            0,
        )
        df = pd.read_csv(io.StringIO("\n".join(lines[header_idx:])), sep="\t", dtype=str)

    df.columns = [str(c).strip() for c in df.columns]
    colmap = {
        "Dispense Time": "dispense_dt",
        "Reception num.": "reception_num",
        "Med ID": "med_id",
        "Medicine Name": "med_desc",
        "Dose Form": "dose_form",
        "Qty per Pack": "qty_per_pack",
        "QOH": "qoh",
        "Manufacturer": "manufacturer",
        "NDC": "ndc",
        "MFG Lot Number": "mfg_lot_number",
        "MFG Expire Date": "mfg_expire_date",
        "Dev.ID": "device_id",
        "Hospital Lot Number": "hospital_lot_number",
        "Hospital Expire Date": "hospital_expire_date",
        "BUD": "bud",
        "User Name": "packaged_by",
        "Confirmer": "confirmer",
    }
    df = df.rename(columns=colmap)

    required = list(colmap.values())
    for col in required:
        if col not in df.columns:
            df[col] = None

    df["dispense_dt"] = pd.to_datetime(df["dispense_dt"], errors="coerce")
    df = df.dropna(subset=["dispense_dt"]).copy()
    df["med_id"] = df["med_id"].fillna("").astype(str).str.strip().str.upper()
    df["med_desc"] = df["med_desc"].fillna("").astype(str).str.strip()
    df["reception_num"] = df["reception_num"].fillna("").astype(str).str.strip()
    df["qty_per_pack"] = pd.to_numeric(df["qty_per_pack"], errors="coerce").fillna(0)
    df["qoh"] = pd.to_numeric(df["qoh"], errors="coerce").fillna(0)
    df["mfg_expire_date"] = pd.to_datetime(df["mfg_expire_date"], errors="coerce").dt.date
    df["hospital_expire_date"] = pd.to_datetime(df["hospital_expire_date"], errors="coerce").dt.date
    df["bud"] = pd.to_datetime(df["bud"], errors="coerce").dt.date

    for col in [
        "dose_form", "manufacturer", "ndc", "mfg_lot_number", "device_id",
        "hospital_lot_number", "packaged_by", "confirmer",
    ]:
        df[col] = df[col].fillna("").astype(str).str.strip()

    df["pk"] = df.apply(
        lambda row: hashlib.sha256(
            "|".join([
                str(row["dispense_dt"]),
                str(row["reception_num"]),
                str(row["med_id"]),
                str(row["hospital_lot_number"]),
            ]).encode()
        ).hexdigest(),
        axis=1,
    )
    return df[["pk"] + required]


def clean_device_inventory(df):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    original_columns = list(df.columns)
    standard_stock_column_v = df[original_columns[21]].copy() if len(original_columns) >= 22 else None

    med_desc_col = "MedDescription.1" if "MedDescription.1" in df.columns else "MedDescription"
    colmap = {
        med_desc_col: "med_desc",
        "Device": "device",
        "Zone": "zone",
        "DrwSubDrwPkt": "pocket_location",
        "Status": "status",
        "BrandName": "brand_name",
        "MedID": "med_id",
        "MedClass": "med_class",
        "CurrentQuantity": "current_quantity",
        "Min": "min_qty",
        "Max": "max_qty",
        "OutdateTracking": "outdate_tracking",
        "LoadedAsFraction": "loaded_as_fraction",
        "Backordered": "backordered",
        "StandardStock": "standard_stock",
        "ActiveOrders": "active_orders",
        "DaysUnused": "days_unused",
    }
    df = df.rename(columns=colmap)

    # Pyxis Device Inventory exports put StandardStock in column V. Use that
    # source when it looks like the expected Y/N flag, even if the header shifted.
    if standard_stock_column_v is not None:
        column_v = standard_stock_column_v.fillna("").astype(str).str.strip().str.upper()
        nonblank_v = column_v[column_v.ne("")]
        column_v_is_yes_no = not nonblank_v.empty and nonblank_v.isin(["Y", "N"]).mean() >= 0.95
        if column_v_is_yes_no:
            df["standard_stock"] = column_v
    elif "standard_stock" not in df.columns:
        df["standard_stock"] = None

    required = list(colmap.values())
    for col in required:
        if col not in df.columns:
            df[col] = None

    for col in [
        "med_desc", "device", "zone", "pocket_location", "status", "brand_name",
        "med_id", "med_class", "outdate_tracking", "loaded_as_fraction",
        "backordered", "standard_stock", "active_orders",
    ]:
        df[col] = df[col].fillna("").astype(str).str.strip()

    df["med_id"] = df["med_id"].str.upper()
    for col in ["current_quantity", "min_qty", "max_qty", "days_unused"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df = df[df["device"].ne("") & df["med_id"].ne("")].copy()
    df["pk"] = df.apply(
        lambda row: hashlib.sha256(
            "|".join([
                row["device"],
                row["med_id"],
                row["pocket_location"],
            ]).encode()
        ).hexdigest(),
        axis=1,
    )
    return df[["pk"] + required]


def clean_iv_room_report(df):
    df = df.copy()
    colmap = {
        "Facility Name": "facility_name",
        "Order/LOT Number": "order_lot_number",
        "Compound Type": "compound_type",
        "# of Preparations": "num_preparations",
        "Dose Number": "dose_number",
        "Drug Name": "drug_name",
        "Order Date": "order_date",
        "Ordered Time": "ordered_time",
        "Completed On": "completed_on",
        "Priority Name": "priority_name",
        "Prepare TAT Minutes": "prepare_tat_minutes",
        "Prepared By": "prepared_by",
        "Approved By": "approved_by",
        "Secondary Approved By": "secondary_approved_by",
    }
    df.rename(columns=colmap, inplace=True)

    required = list(colmap.values())
    for col in required:
        if col not in df.columns:
            df[col] = None

    for col in ["facility_name", "compound_type", "drug_name",
                "ordered_time", "priority_name", "prepared_by", "approved_by", "secondary_approved_by"]:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace({"": None, "nan": None, "None": None})

    df["order_lot_number"] = df["order_lot_number"].apply(normalize_identifier_text)
    df["dose_number"] = df["dose_number"].apply(normalize_identifier_text)

    df["order_date"] = pd.to_datetime(df["order_date"], errors="coerce").dt.date
    df["ordered_time"] = df["ordered_time"].astype(str).str.strip()
    df["ordered_time"] = df["ordered_time"].replace({"": None, "nan": None, "None": None})
    df["order_dt"] = pd.to_datetime(
        df["order_date"].astype(str) + " " + df["ordered_time"].fillna("00:00"),
        errors="coerce",
    )

    df["completed_on"] = pd.to_datetime(df["completed_on"], errors="coerce")
    df["num_preparations"] = pd.to_numeric(df["num_preparations"], errors="coerce").fillna(1)
    df["prepare_tat_minutes"] = pd.to_numeric(df["prepare_tat_minutes"], errors="coerce")

    df.dropna(subset=["order_date"], inplace=True)
    df["pk"] = df.apply(
        lambda row: hashlib.sha256(
            "|".join(
                [
                    str(row.get("facility_name") or ""),
                    str(row.get("order_lot_number") or ""),
                    str(row.get("drug_name") or ""),
                    str(row.get("order_date") or ""),
                    str(row.get("ordered_time") or ""),
                    str(row.get("dose_number") or ""),
                ]
            ).encode()
        ).hexdigest(),
        axis=1,
    )

    df = df.astype(object)
    df = df.where(pd.notna(df), None)
    return df[required + ["order_dt", "pk"]]


def clean_iv_room_workflow_detail(df, source_file=""):
    df = df.copy()
    colmap = {
        "Facility Name": "facility_name",
        "Order/Lot Number": "order_lot_number",
        "Order/LOT Number": "order_lot_number",
        "Dose Number": "dose_number",
        "Ordered On": "ordered_on",
        "Prepared By": "prepared_by",
        "Approved By": "approved_by",
        "Drug Name": "drug_name",
        "Workflow Name": "workflow_name",
        "Workflow Step Type": "workflow_step_type",
        "Workflow Step Name": "workflow_step_name",
        "Workflow Step Category": "workflow_step_category",
        "Start Date": "start_date",
        "Start Time": "start_time",
        "Stop Time": "stop_time",
        "Total Duration Minutes": "total_duration_minutes",
    }
    df.rename(columns=colmap, inplace=True)

    required = list(dict.fromkeys(colmap.values()))
    for col in required:
        if col not in df.columns:
            df[col] = None

    for col in [
        "facility_name", "prepared_by", "approved_by", "drug_name", "workflow_name",
        "workflow_step_type", "workflow_step_name", "workflow_step_category",
        "start_time", "stop_time",
    ]:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace({"": None, "nan": None, "None": None})

    df["order_lot_number"] = df["order_lot_number"].apply(normalize_identifier_text)
    df["dose_number"] = df["dose_number"].apply(normalize_identifier_text)
    df["ordered_on"] = pd.to_datetime(df["ordered_on"], errors="coerce").dt.date
    df["start_date"] = pd.to_datetime(df["start_date"], errors="coerce").dt.date
    df["total_duration_minutes"] = pd.to_numeric(df["total_duration_minutes"], errors="coerce")
    df["source_file"] = source_file

    df["start_dt"] = pd.to_datetime(
        df["start_date"].astype(str) + " " + df["start_time"].fillna("00:00"),
        errors="coerce",
    )
    df["stop_dt"] = pd.NaT
    has_stop_time = df["stop_time"].notna()
    df.loc[has_stop_time, "stop_dt"] = pd.to_datetime(
        df.loc[has_stop_time, "start_date"].astype(str) + " " + df.loc[has_stop_time, "stop_time"],
        errors="coerce",
    )

    df.dropna(subset=["start_date"], inplace=True)
    df["pk"] = df.apply(
        lambda row: hashlib.sha256(
            "|".join([
                str(row.get("order_lot_number") or ""),
                str(row.get("dose_number") or ""),
                str(row.get("drug_name") or ""),
                str(row.get("workflow_name") or ""),
                str(row.get("workflow_step_type") or ""),
                str(row.get("workflow_step_name") or ""),
                str(row.get("workflow_step_category") or ""),
                str(row.get("start_date") or ""),
                str(row.get("start_time") or ""),
                str(row.get("stop_time") or ""),
                str(row.get("total_duration_minutes") or ""),
            ]).encode()
        ).hexdigest(),
        axis=1,
    )

    df = df.astype(object)
    df = df.where(pd.notna(df), None)
    return df[[
        "pk", "facility_name", "order_lot_number", "dose_number", "ordered_on",
        "prepared_by", "approved_by", "drug_name", "workflow_name",
        "workflow_step_type", "workflow_step_name", "workflow_step_category",
        "start_date", "start_time", "stop_time", "start_dt", "stop_dt",
        "total_duration_minutes", "source_file",
    ]]


def clean_wcc_compounding_stats(df, source_file=""):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    def _norm_col(value):
        return re.sub(r"[^a-z0-9]+", "", str(value).lower())

    normalized_cols = {_norm_col(col): col for col in df.columns}

    def _pick_col(*aliases):
        for alias in aliases:
            match = normalized_cols.get(_norm_col(alias))
            if match is not None:
                return match
        alias_tokens = [_norm_col(alias) for alias in aliases]
        for norm_name, original_name in normalized_cols.items():
            if any(token and token in norm_name for token in alias_tokens):
                return original_name
        return None

    component_col = _pick_col("Slices by Medication Component Name", "Medication Component Name", "Component Name")
    order_col = _pick_col("Order Name", "Order")
    admin_col = _pick_col("Administration Instant", "Administration Date Time", "Administration Date & Time", "Admin Instant")
    barcode_col = _pick_col("Barcode Scanning Compliance Status", "Barcode Compliance Status", "Barcode Status")

    if not component_col and len(df.columns) >= 4:
        component_col, order_col, admin_col, barcode_col = df.columns[:4]

    def _series_or_default(col, default):
        if col in df.columns:
            return df[col]
        return pd.Series([default] * len(df), index=df.index)

    clean_df = pd.DataFrame({
        "component_name": _series_or_default(component_col, ""),
        "order_name": _series_or_default(order_col, ""),
        "administration_dt": _series_or_default(admin_col, None),
        "barcode_status": _series_or_default(barcode_col, ""),
    })

    for col in ["component_name", "order_name", "barcode_status"]:
        clean_df[col] = clean_df[col].fillna("").astype(str).str.strip()

    component_id = clean_df["component_name"].str.extract(r"\[([^\]]+)\]", expand=False)
    clean_df["component_id"] = component_id.fillna("").astype(str).str.strip()

    clean_df["administration_dt"] = parse_excel_datetime_series(clean_df["administration_dt"])

    df = clean_df[clean_df["component_name"].ne("") & clean_df["administration_dt"].notna()].copy()
    df["source_file"] = source_file or ""
    df["pk"] = df.apply(
        lambda row: hashlib.sha256(
            "|".join([
                str(row.get("component_name") or ""),
                str(row.get("order_name") or ""),
                str(row.get("administration_dt") or ""),
                str(row.get("barcode_status") or ""),
            ]).encode()
        ).hexdigest(),
        axis=1,
    )
    df = df.astype(object)
    df = df.where(pd.notna(df), None)
    return df[["pk", "component_name", "component_id", "order_name", "administration_dt", "barcode_status", "source_file"]]


def parse_excel_datetime_series(series):
    numeric_values = pd.to_numeric(series, errors="coerce")
    parsed_values = pd.to_datetime(series.where(numeric_values.isna()), errors="coerce")
    excel_values = pd.to_datetime(numeric_values, unit="D", origin="1899-12-30", errors="coerce")
    return excel_values.where(numeric_values.notna(), parsed_values)


def infer_cartfill_area(order_medication, pharmacy, location):
    text = " ".join([
        str(order_medication or ""),
        str(pharmacy or ""),
        str(location or ""),
    ]).upper()
    if any(token in text for token in ["WCC", "WOMEN", "WOMENS", "CHILD", "CHILDREN", "CANCER"]):
        return "WCC"
    if any(token in text for token in ["CLEANROOM", "IV ROOM", "STERILE"]):
        return "IV Room"
    if any(token in text for token in ["CAROUSEL", "CENTRAL PHARMACY", "SJS CENTRAL", "MAIN PHARMACY"]):
        return "Central Pharmacy"
    return "Needs Review"


def clean_wcc_cartfill_stats(df, source_file=""):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    colmap = {
        "Start Date": "report_start_date",
        "End Date": "report_end_date",
        "Order Medication": "order_medication",
        "Ready for Dispense for Date & Time": "ready_for_dispense_dt",
        "Admin Given Date & Time": "admin_given_dt",
        "Prepared Date & Time": "prepared_dt",
        "Prep or Dispense User": "prep_or_dispense_user",
        "Location": "location",
        "Pharmacy": "pharmacy",
    }
    df.rename(columns=colmap, inplace=True)

    required = list(colmap.values())
    for col in required:
        if col not in df.columns:
            df[col] = None

    for col in ["order_medication", "prep_or_dispense_user", "location", "pharmacy"]:
        df[col] = df[col].fillna("").astype(str).str.strip()

    df["med_id"] = df["order_medication"].str.extract(r"\[([^\]]+)\]", expand=False).fillna("").astype(str).str.strip()
    df["report_start_date"] = parse_excel_datetime_series(df["report_start_date"]).dt.date
    df["report_end_date"] = parse_excel_datetime_series(df["report_end_date"]).dt.date
    df["ready_for_dispense_dt"] = parse_excel_datetime_series(df["ready_for_dispense_dt"])
    df["admin_given_dt"] = parse_excel_datetime_series(df["admin_given_dt"])
    df["prepared_dt"] = parse_excel_datetime_series(df["prepared_dt"])
    df["cartfill_area"] = df.apply(
        lambda row: infer_cartfill_area(row.get("order_medication"), row.get("pharmacy"), row.get("location")),
        axis=1,
    )

    df = df[df["order_medication"].ne("") & df["ready_for_dispense_dt"].notna()].copy()
    df["source_file"] = source_file or ""
    df["pk"] = df.apply(
        lambda row: hashlib.sha256(
            "|".join([
                str(row.get("order_medication") or ""),
                str(row.get("ready_for_dispense_dt") or ""),
                str(row.get("prepared_dt") or ""),
                str(row.get("pharmacy") or ""),
                str(row.get("location") or ""),
            ]).encode()
        ).hexdigest(),
        axis=1,
    )
    df = df.astype(object)
    df = df.where(pd.notna(df), None)
    return df[[
        "pk", "report_start_date", "report_end_date", "order_medication", "med_id",
        "ready_for_dispense_dt", "admin_given_dt", "prepared_dt", "prep_or_dispense_user",
        "location", "pharmacy", "cartfill_area", "source_file",
    ]]


def clean_overnight_cartfill_workbook(uploaded):
    uploaded.seek(0)
    excel = pd.ExcelFile(uploaded)

    orders = pd.DataFrame()
    windows = pd.DataFrame()
    staffing = pd.DataFrame()

    if "Cartfill Data" in excel.sheet_names:
        raw_orders = pd.read_excel(excel, sheet_name="Cartfill Data")
        raw_orders = raw_orders.rename(columns={
            "Order ID": "order_id",
            "Order Medication": "order_medication",
            "Ready for Dispense for Date & Time": "ready_for_dispense_dt",
            "Admin Given Date & Time": "admin_given_dt",
            "Prepared Date & Time": "prepared_dt",
            "Prep or Dispense User": "prep_or_dispense_user",
            "Pharmacy": "pharmacy",
        }).copy()

        required_cols = [
            "order_id", "order_medication", "ready_for_dispense_dt", "admin_given_dt",
            "prepared_dt", "prep_or_dispense_user", "pharmacy",
        ]
        for col in required_cols:
            if col not in raw_orders.columns:
                raw_orders[col] = None

        for col in ["ready_for_dispense_dt", "admin_given_dt", "prepared_dt"]:
            raw_orders[col] = excel_serial_to_datetime(raw_orders[col])

        raw_orders["pharmacy"] = raw_orders["pharmacy"].fillna("").astype(str).str.strip()
        raw_orders["prep_or_dispense_user"] = raw_orders["prep_or_dispense_user"].fillna("").astype(str).str.strip()
        raw_orders["order_medication"] = raw_orders["order_medication"].fillna("").astype(str).str.strip()
        raw_orders["order_id"] = raw_orders["order_id"].apply(normalize_identifier_text)
        raw_orders["order_id"] = raw_orders["order_id"].fillna("")

        raw_orders["event_date"] = (
            raw_orders["admin_given_dt"]
            .fillna(raw_orders["ready_for_dispense_dt"])
            .fillna(raw_orders["prepared_dt"])
            .dt.date
        )
        raw_orders["required_start_dt"] = raw_orders["ready_for_dispense_dt"] - pd.Timedelta(hours=4)
        raw_orders["prep_lead_hours"] = (
            raw_orders["ready_for_dispense_dt"] - raw_orders["prepared_dt"]
        ).dt.total_seconds() / 3600
        raw_orders["hold_hours"] = (
            raw_orders["admin_given_dt"] - raw_orders["prepared_dt"]
        ).dt.total_seconds() / 3600
        raw_orders["is_sjs_cleanroom"] = raw_orders["pharmacy"].str.upper().eq("SJS CLEANROOM")

        raw_orders = raw_orders[
            raw_orders["event_date"].notna() &
            raw_orders["order_id"].ne("")
        ].copy()

        raw_orders["pk"] = raw_orders.apply(
            lambda row: hashlib.sha256(
                "|".join([
                    str(row.get("order_id") or ""),
                    str(row.get("ready_for_dispense_dt") or ""),
                    str(row.get("admin_given_dt") or ""),
                    str(row.get("prepared_dt") or ""),
                    str(row.get("pharmacy") or ""),
                ]).encode()
            ).hexdigest(),
            axis=1,
        )
        raw_orders = raw_orders.astype(object).where(pd.notna(raw_orders), None)
        orders = raw_orders[[
            "pk", "order_id", "order_medication", "ready_for_dispense_dt",
            "admin_given_dt", "prepared_dt", "prep_or_dispense_user", "pharmacy",
            "event_date", "required_start_dt", "prep_lead_hours", "hold_hours",
            "is_sjs_cleanroom",
        ]]

    if "Cartfill times 04.2026" in excel.sheet_names:
        raw_windows = pd.read_excel(excel, sheet_name="Cartfill times 04.2026")
        raw_windows = raw_windows.rename(columns={
            "Cartfill Name": "cartfill_name",
            "Time Processed": "time_processed_raw",
            "Doses Due": "doses_due",
            "Pharmacy": "pharmacy",
        }).copy()
        for col in ["cartfill_name", "time_processed_raw", "doses_due", "pharmacy"]:
            if col not in raw_windows.columns:
                raw_windows[col] = None
            raw_windows[col] = raw_windows[col].fillna("").astype(str).str.strip()

        raw_windows = raw_windows[raw_windows["cartfill_name"].ne("")].copy()
        raw_windows["pk"] = raw_windows.apply(
            lambda row: hashlib.sha256(
                "|".join([
                    str(row.get("cartfill_name") or ""),
                    str(row.get("time_processed_raw") or ""),
                    str(row.get("doses_due") or ""),
                    str(row.get("pharmacy") or ""),
                ]).encode()
            ).hexdigest(),
            axis=1,
        )
        raw_windows = raw_windows.astype(object).where(pd.notna(raw_windows), None)
        windows = raw_windows[["pk", "cartfill_name", "time_processed_raw", "doses_due", "pharmacy"]]

    if "Schedule" in excel.sheet_names:
        raw_schedule = pd.read_excel(excel, sheet_name="Schedule", header=None)
        shift_headers = raw_schedule.iloc[0].tolist() if len(raw_schedule.index) > 0 else []
        weekend_headers = raw_schedule.iloc[1].tolist() if len(raw_schedule.index) > 1 else []
        rows = []
        for idx in range(2, len(raw_schedule.index)):
            row = raw_schedule.iloc[idx]
            schedule_date = excel_serial_to_datetime(pd.Series([row.iloc[1]])).iloc[0]
            day_name = None if pd.isna(row.iloc[2]) else str(row.iloc[2]).strip()
            if pd.isna(schedule_date) and not day_name:
                continue
            for col_idx in range(3, min(len(row), len(shift_headers))):
                shift_name = shift_headers[col_idx] if col_idx < len(shift_headers) else None
                assigned_staff = row.iloc[col_idx]
                if pd.isna(shift_name) or pd.isna(assigned_staff):
                    continue
                assigned_staff = str(assigned_staff).strip()
                if not assigned_staff:
                    continue
                weekend_label = weekend_headers[col_idx] if col_idx < len(weekend_headers) else None
                rows.append({
                    "schedule_date": schedule_date.date() if pd.notna(schedule_date) else None,
                    "day_name": day_name,
                    "shift_name": str(shift_name).strip(),
                    "weekend_shift_label": None if pd.isna(weekend_label) else str(weekend_label).strip(),
                    "assigned_staff": assigned_staff,
                    "is_weekend": str(day_name).lower() in {"saturday", "sunday"} if day_name else False,
                    "is_placeholder": assigned_staff.lower() in {"x", "w"},
                })
        if rows:
            staffing = pd.DataFrame(rows)
            staffing["pk"] = staffing.apply(
                lambda row: hashlib.sha256(
                    "|".join([
                        str(row.get("schedule_date") or ""),
                        str(row.get("day_name") or ""),
                        str(row.get("shift_name") or ""),
                        str(row.get("assigned_staff") or ""),
                    ]).encode()
                ).hexdigest(),
                axis=1,
            )
            staffing = staffing.astype(object).where(pd.notna(staffing), None)
            staffing = staffing[[
                "pk", "schedule_date", "day_name", "shift_name", "weekend_shift_label",
                "assigned_staff", "is_weekend", "is_placeholder",
            ]]

    return {"orders": orders, "windows": windows, "staffing": staffing}

# --- DATA LOADERS (CACHED) ---
@st.cache_data(ttl=300)
def load_data(start_date, end_date):
    queries = {
        "events": """
            WITH audit_days AS (
                SELECT DISTINCT dt::date AS d
                FROM audit_transaction_detail_rc
                WHERE dt::date BETWEEN %s AND %s
            ),
            audit_events AS (
                SELECT
                    a.user_name,
                    a.station_name AS device,
                    a.med_id,
                    a.med_desc,
                    a.transaction_type AS event_type,
                    a.dt,
                    a.qty,
                    a.beginning_qty,
                    a.ending_qty,
                    a.discrepancy_difference AS discrepancy_qty,
                    COALESCE(NULLIF(a.discrepancy_reason, ''), a.discrepancy_resolution_desc) AS discrepancy_reason,
                    COALESCE(a.unit_cost, c.cost_per_unit, 0) AS cost_per_unit,
                    a.pk
                FROM audit_transaction_detail_rc a
                LEFT JOIN med_costs c ON UPPER(TRIM(a.med_id)) = UPPER(TRIM(c.med_id))
                WHERE a.dt::date BETWEEN %s AND %s
            ),
            legacy_events AS (
                SELECT
                    e.user_name,
                    e.device,
                    e.med_id,
                    e.med_desc,
                    e.event_type,
                    e.dt,
                    e.qty,
                    e.beginning_qty,
                    e.ending_qty,
                    e.discrepancy_qty,
                    e.discrepancy_reason,
                    COALESCE(c.cost_per_unit, 0) AS cost_per_unit,
                    e.pk
                FROM events e
                LEFT JOIN med_costs c ON UPPER(TRIM(e.med_id)) = UPPER(TRIM(c.med_id))
                WHERE e.dt::date BETWEEN %s AND %s
                  AND NOT EXISTS (
                      SELECT 1
                      FROM audit_days ad
                      WHERE ad.d = e.dt::date
                  )
            )
            SELECT * FROM audit_events
            UNION ALL
            SELECT * FROM legacy_events
        """,
        "config": """
            SELECT pk, dt, user_name, device, med_id, location, action_type, activity_category, 
                   min_qty, max_qty, is_standard 
            FROM config_events WHERE dt::date BETWEEN %s AND %s
        """,
        "pharm": """
            SELECT pk, queue_id, priority, dt, med_id, med_desc, destination, user_name, qty
            FROM pharmacy_orders WHERE dt::date BETWEEN %s AND %s
        """,
        "schedule": """
            SELECT pk, dt, day_name, staff_name, shift_type, assignment_type, note,
                   COALESCE(schedule_status, assignment_type, 'Standard') AS schedule_status,
                   cell_fill_color
            FROM staff_schedule WHERE dt BETWEEN %s AND %s
        """,
        "attendance": """
            SELECT pk, raw_name, dt_date, start_dt, end_dt
            FROM attendance_punches WHERE dt_date BETWEEN %s AND %s
        """
    }
    
    results = {}
    params = (start_date, end_date)
    with db_cursor() as (conn, cur):
        for key, sql in queries.items():
            try:
                query_params = params * 3 if key == "events" else params
                results[key] = pd.read_sql(sql, conn, params=query_params)
                if not results[key].empty and 'dt' in results[key].columns:
                    results[key]["dt"] = pd.to_datetime(results[key]["dt"])
            except Exception:
                results[key] = pd.DataFrame()

    df = results["events"]
    if not df.empty:

    # 🔒 GLOBAL NONE GUARD (CRITICAL)
        df["user_name"] = (
            df["user_name"]
            .fillna("unknown")
            .astype(str)
            .str.strip()
            .replace(["None", "none", ""], "unknown")
        )
    
        df["med_desc"] = df["med_desc"].fillna("unknown").astype(str)
        df["device"] = df["device"].fillna("unknown").astype(str)
    
        # Numeric stability
        df["cost_per_unit"] = df["cost_per_unit"].fillna(0).astype('float32')
        df["qty"] = df["qty"].fillna(0).astype('float32')
        for qty_col in ["beginning_qty", "ending_qty", "discrepancy_qty"]:
            if qty_col in df.columns:
                df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0).astype("float32")

    
        df = df[~df['med_desc'].astype(str).str.contains(r'Drw|Pkt|Cubic', regex=True, case=False, na=False)]
        df.sort_values(['user_name', 'dt'], inplace=True)
        df['next_dt'] = df.groupby('user_name')['dt'].shift(-1)
        df['duration'] = (df['next_dt'] - df['dt']).dt.total_seconds()
        df['prev_device'] = df.groupby('user_name')['device'].shift(1)
        df['gap_prev'] = (df['dt'] - df.groupby('user_name')['dt'].shift(1)).dt.total_seconds().fillna(0)
        df['machine_time_sec'] = np.where((df['device'] == df.groupby('user_name')['device'].shift(-1)) & (df['duration'] < 600), df['duration'], 0)
        df['is_new_session'] = np.where((df['user_name'] != df['user_name'].shift(1)) | (df['device'] != df['prev_device']) | (df['gap_prev'] > 1200), 1, 0)
        df['session_id'] = df['is_new_session'].cumsum()
        df.drop(columns=['next_dt', 'is_new_session', 'gap_prev'], inplace=True, errors='ignore')

    if not results["pharm"].empty:
        results["pharm"] = results["pharm"][~results["pharm"]['destination'].astype(str).str.contains('BATCH PICK', case=False, na=False)]

    return df, results["config"], results["pharm"], results["schedule"], results["attendance"]


@st.cache_data(ttl=300)
def load_pharmacy_workflow_orders(start_date, end_date):
    sql = """
        SELECT pk, queue_id, priority, dt, med_id, med_desc, destination, user_name, qty
        FROM pharmacy_orders
        WHERE dt::date BETWEEN %s AND %s
    """
    df = run_query(sql, params=(start_date, end_date))
    if df.empty:
        return df
    df["dt"] = pd.to_datetime(df["dt"], errors="coerce")
    for col in ["priority", "med_id", "med_desc", "destination", "user_name"]:
        df[col] = df[col].fillna("").astype(str).str.strip()
    df["qty"] = pd.to_numeric(df["qty"], errors="coerce").fillna(0)
    return df[~df["destination"].str.contains("BATCH PICK", case=False, na=False)].copy()


@st.cache_data(ttl=300)
def load_pyxis_workflow_events(start_date, end_date):
    sql = """
        WITH audit_days AS (
            SELECT DISTINCT dt::date AS d
            FROM audit_transaction_detail_rc
            WHERE dt::date BETWEEN %s AND %s
        ),
        audit_events AS (
            SELECT
                a.user_name,
                a.station_name AS device,
                a.med_id,
                a.med_desc,
                a.transaction_type AS event_type,
                a.dt,
                a.qty,
                a.discrepancy_difference AS discrepancy_qty,
                COALESCE(NULLIF(a.discrepancy_reason, ''), a.discrepancy_resolution_desc) AS discrepancy_reason,
                a.ending_qty,
                COALESCE(a.unit_cost, c.cost_per_unit, 0) AS cost_per_unit
            FROM audit_transaction_detail_rc a
            LEFT JOIN med_costs c ON UPPER(TRIM(a.med_id)) = UPPER(TRIM(c.med_id))
            WHERE a.dt::date BETWEEN %s AND %s
        ),
        legacy_events AS (
            SELECT
                e.user_name,
                e.device,
                e.med_id,
                e.med_desc,
                e.event_type,
                e.dt,
                e.qty,
                e.discrepancy_qty,
                e.discrepancy_reason,
                e.ending_qty,
                COALESCE(c.cost_per_unit, 0) AS cost_per_unit
            FROM events e
            LEFT JOIN med_costs c ON UPPER(TRIM(e.med_id)) = UPPER(TRIM(c.med_id))
            WHERE e.dt::date BETWEEN %s AND %s
              AND NOT EXISTS (
                  SELECT 1
                  FROM audit_days ad
                  WHERE ad.d = e.dt::date
              )
        )
        SELECT * FROM audit_events
        UNION ALL
        SELECT * FROM legacy_events
    """
    df = run_query(sql, params=(start_date, end_date) * 3)
    if df.empty:
        return df
    df["dt"] = pd.to_datetime(df["dt"], errors="coerce")
    for col in ["user_name", "device", "med_id", "med_desc", "event_type", "discrepancy_reason"]:
        df[col] = df[col].fillna("").astype(str).str.strip()
    for col in ["qty", "discrepancy_qty", "ending_qty", "cost_per_unit"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def get_zero_verify_events(df_events):
    if df_events.empty or not {"event_type", "qty"}.issubset(df_events.columns):
        return pd.DataFrame()

    zero_events = df_events.copy()
    zero_events["_etype"] = zero_events["event_type"].astype(str).str.lower().str.strip()
    zero_events["_qty"] = pd.to_numeric(zero_events["qty"], errors="coerce")
    zero_events = zero_events[
        zero_events["_etype"].str.contains("verify", na=False) &
        zero_events["_qty"].eq(0)
    ].copy()

    if zero_events.empty:
        return zero_events

    available_cols = [
        "dt", "user_name", "device", "med_id", "med_desc", "event_type",
        "qty", "beginning_qty", "ending_qty", "discrepancy_qty", "discrepancy_reason", "pk",
    ]
    available_cols = [col for col in available_cols if col in zero_events.columns]
    return zero_events.sort_values("dt", ascending=False)[available_cols]


def summarize_zero_verify_events(zero_events):
    if zero_events.empty:
        return pd.DataFrame()

    summary = (
        zero_events
        .sort_values("dt")
        .groupby(["device", "med_id", "med_desc"], dropna=False)
        .agg(
            zero_verifies=("pk", "count"),
            first_seen=("dt", "min"),
            last_seen=("dt", "max"),
            verify_users=("user_name", lambda s: ", ".join(sorted({str(v) for v in s.dropna()}))),
        )
        .reset_index()
        .sort_values(["last_seen", "zero_verifies"], ascending=[False, False])
    )
    return summary


@st.cache_data(ttl=300)
def load_zero_verify_refill_gaps(start_date, end_date, lookback_days=60):
    lookback_start = start_date - timedelta(days=lookback_days)
    sql = text("""
        WITH zero_verifies AS (
            SELECT
                pk,
                dt::timestamp AS dt,
                user_name,
                device,
                med_id,
                med_desc,
                event_type,
                qty,
                beginning_qty,
                ending_qty,
                discrepancy_qty
            FROM events
            WHERE dt::date BETWEEN :start_date AND :end_date
              AND event_type ILIKE '%verify%'
              AND COALESCE(qty, 0) = 0
        ),
        refill_events AS (
            SELECT
                dt::timestamp AS dt,
                user_name,
                device,
                med_id,
                event_type,
                qty
            FROM events
            WHERE dt::date BETWEEN :lookback_start AND :end_date
              AND event_type ILIKE ANY (ARRAY['%restock%', '%refill%', '%load%', '%replenish%'])
              AND event_type NOT ILIKE '%cancel%'
              AND event_type NOT ILIKE '%unload%'
              AND event_type NOT ILIKE '%empty%'
        )
        SELECT
            z.pk,
            z.dt,
            z.user_name,
            z.device,
            z.med_id,
            z.med_desc,
            z.event_type,
            z.qty,
            z.beginning_qty,
            z.ending_qty,
            z.discrepancy_qty,
            r.dt AS prior_refill_dt,
            r.user_name AS prior_refill_user,
            r.event_type AS prior_refill_event_type,
            r.qty AS prior_refill_qty,
            EXTRACT(EPOCH FROM (z.dt - r.dt)) / 3600.0 AS hours_since_refill
        FROM zero_verifies z
        LEFT JOIN LATERAL (
            SELECT dt, user_name, event_type, qty
            FROM refill_events r
            WHERE r.dt < z.dt
              AND COALESCE(r.device, '') = COALESCE(z.device, '')
              AND COALESCE(r.med_id, '') = COALESCE(z.med_id, '')
            ORDER BY r.dt DESC
            LIMIT 1
        ) r ON TRUE
        ORDER BY z.dt DESC
    """)
    try:
        with engine.connect() as conn:
            df = pd.read_sql(
                sql,
                conn,
                params={
                    "lookback_start": str(lookback_start),
                    "start_date": str(start_date),
                    "end_date": str(end_date),
                },
            )
        if df.empty:
            return df
        for col in ["dt", "prior_refill_dt"]:
            df[col] = pd.to_datetime(df[col], errors="coerce")
        for col in ["qty", "beginning_qty", "ending_qty", "discrepancy_qty", "prior_refill_qty", "hours_since_refill"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        for col in ["user_name", "device", "med_id", "med_desc", "event_type", "prior_refill_user", "prior_refill_event_type"]:
            if col in df.columns:
                df[col] = df[col].fillna("").astype(str).str.strip()
        return df
    except Exception as exc:
        st.warning(f"Could not load zero-verify refill gaps: {exc}")
        return pd.DataFrame()


def summarize_zero_verify_refill_gaps(zero_gap_events):
    if zero_gap_events.empty:
        return pd.DataFrame()

    summary = (
        zero_gap_events
        .sort_values("dt")
        .groupby(["device", "med_id", "med_desc"], dropna=False)
        .agg(
            zero_verifies=("pk", "count"),
            first_seen=("dt", "min"),
            last_seen=("dt", "max"),
            avg_hours_since_refill=("hours_since_refill", "mean"),
            median_hours_since_refill=("hours_since_refill", "median"),
            last_prior_refill=("prior_refill_dt", "max"),
            verify_users=("user_name", lambda s: ", ".join(sorted({str(v) for v in s.dropna() if str(v)}))),
            prior_refill_users=("prior_refill_user", lambda s: ", ".join(sorted({str(v) for v in s.dropna() if str(v)}))),
        )
        .reset_index()
        .sort_values(["last_seen", "zero_verifies"], ascending=[False, False])
    )
    return summary


@st.cache_data(ttl=300)
def load_iv_room_data(start_date, end_date):
    query = """
        SELECT
            pk,
            facility_name,
            order_lot_number,
            compound_type,
            num_preparations,
            dose_number,
            drug_name,
            order_date,
            ordered_time,
            order_dt,
            completed_on,
            priority_name,
            prepare_tat_minutes,
            prepared_by,
            approved_by,
            secondary_approved_by
        FROM iv_room_workload
        WHERE order_date BETWEEN %s AND %s
        ORDER BY order_dt, order_lot_number
    """
    try:
        with db_cursor() as (conn, cur):
            df = pd.read_sql(query, conn, params=(start_date, end_date))
        for col in ["order_date", "order_dt", "completed_on"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")
        if "order_lot_number" in df.columns:
            df["order_lot_number"] = df["order_lot_number"].apply(normalize_identifier_text)
        if "dose_number" in df.columns:
            df["dose_number"] = df["dose_number"].apply(normalize_identifier_text)
        dedupe_cols = [
            c for c in [
                "facility_name", "order_lot_number", "compound_type", "num_preparations",
                "dose_number", "drug_name", "order_date", "ordered_time", "order_dt",
                "completed_on", "priority_name", "prepared_by", "approved_by", "secondary_approved_by"
            ] if c in df.columns
        ]
        if dedupe_cols:
            df = df.drop_duplicates(subset=dedupe_cols, keep="first").copy()
        return df
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=300)
def load_iv_room_workflow_detail(start_date, end_date):
    query = """
        SELECT
            pk,
            facility_name,
            order_lot_number,
            dose_number,
            ordered_on,
            prepared_by,
            approved_by,
            drug_name,
            workflow_name,
            workflow_step_type,
            workflow_step_name,
            workflow_step_category,
            start_date,
            start_time,
            stop_time,
            start_dt,
            stop_dt,
            total_duration_minutes,
            source_file
        FROM iv_room_workflow_detail
        WHERE start_date BETWEEN %s AND %s
        ORDER BY start_dt, order_lot_number, dose_number, workflow_step_type, workflow_step_name
    """
    try:
        with db_cursor() as (conn, cur):
            df = pd.read_sql(query, conn, params=(start_date, end_date))
        for col in ["ordered_on", "start_date", "start_dt", "stop_dt"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")
        for col in ["order_lot_number", "dose_number"]:
            if col in df.columns:
                df[col] = df[col].apply(normalize_identifier_text)
        if "total_duration_minutes" in df.columns:
            df["total_duration_minutes"] = pd.to_numeric(df["total_duration_minutes"], errors="coerce")
        return df
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=300)
def load_wcc_compounding_stats(start_date, end_date):
    query = """
        SELECT
            pk,
            component_name,
            component_id,
            order_name,
            administration_dt,
            barcode_status,
            source_file,
            uploaded_at
        FROM wcc_compounding_stats
        WHERE administration_dt::date BETWEEN %s AND %s
        ORDER BY administration_dt
    """
    try:
        with db_cursor() as (conn, cur):
            df = pd.read_sql(query, conn, params=(start_date, end_date))
        for col in ["administration_dt", "uploaded_at"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")
        return df
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=300)
def load_wcc_cartfill_stats(start_date, end_date):
    init_db()
    query = """
        SELECT
            pk,
            report_start_date,
            report_end_date,
            order_medication,
            med_id,
            ready_for_dispense_dt,
            admin_given_dt,
            prepared_dt,
            prep_or_dispense_user,
            location,
            pharmacy,
            COALESCE(cartfill_area, 'Needs Review') AS cartfill_area,
            source_file,
            uploaded_at
        FROM wcc_cartfill_stats
        WHERE ready_for_dispense_dt::date BETWEEN %s AND %s
        ORDER BY ready_for_dispense_dt
    """
    try:
        with db_cursor() as (conn, cur):
            df = pd.read_sql(query, conn, params=(start_date, end_date))
        for col in ["report_start_date", "report_end_date", "ready_for_dispense_dt", "admin_given_dt", "prepared_dt", "uploaded_at"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")
        return df
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=300)
def load_overnight_cartfill_orders(start_date, end_date):
    init_db()
    query = """
        SELECT * FROM (
            SELECT
                'legacy-' || pk AS pk,
                order_id,
                order_medication,
                ready_for_dispense_dt,
                admin_given_dt,
                prepared_dt,
                prep_or_dispense_user,
                pharmacy,
                event_date,
                required_start_dt,
                prep_lead_hours,
                hold_hours,
                is_sjs_cleanroom
            FROM overnight_iv_cartfill_orders
            WHERE event_date BETWEEN %s AND %s

            UNION ALL

            SELECT
                'all-' || pk AS pk,
                pk AS order_id,
                order_medication,
                ready_for_dispense_dt,
                admin_given_dt,
                prepared_dt,
                prep_or_dispense_user,
                pharmacy,
                ready_for_dispense_dt::date AS event_date,
                ready_for_dispense_dt - INTERVAL '4 hours' AS required_start_dt,
                EXTRACT(EPOCH FROM (ready_for_dispense_dt - prepared_dt)) / 3600.0 AS prep_lead_hours,
                NULL::float AS hold_hours,
                (COALESCE(cartfill_area, '') = 'IV Room' OR UPPER(COALESCE(pharmacy, '')) LIKE '%%CLEANROOM%%') AS is_sjs_cleanroom
            FROM wcc_cartfill_stats
            WHERE ready_for_dispense_dt::date BETWEEN %s AND %s
        ) cartfill_orders
        ORDER BY COALESCE(admin_given_dt, ready_for_dispense_dt, prepared_dt)
    """
    try:
        with db_cursor() as (conn, cur):
            df = pd.read_sql(query, conn, params=(start_date, end_date, start_date, end_date))
        for col in ["ready_for_dispense_dt", "admin_given_dt", "prepared_dt", "required_start_dt", "event_date"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")
        if "order_id" in df.columns:
            df["order_id"] = df["order_id"].apply(normalize_identifier_text)
        dedupe_cols = [
            c for c in [
                "order_medication", "ready_for_dispense_dt", "prepared_dt",
                "prep_or_dispense_user", "pharmacy", "event_date", "is_sjs_cleanroom"
            ] if c in df.columns
        ]
        if dedupe_cols:
            df = df.drop_duplicates(subset=dedupe_cols, keep="first").copy()
        return df
    except Exception:
        fallback_query = """
            SELECT
            pk,
            order_id,
            order_medication,
            ready_for_dispense_dt,
            admin_given_dt,
            prepared_dt,
            prep_or_dispense_user,
            pharmacy,
            event_date,
            required_start_dt,
            prep_lead_hours,
            hold_hours,
            is_sjs_cleanroom
        FROM overnight_iv_cartfill_orders
        WHERE event_date BETWEEN %s AND %s
        ORDER BY COALESCE(admin_given_dt, ready_for_dispense_dt, prepared_dt)
    """
        with db_cursor() as (conn, cur):
            df = pd.read_sql(fallback_query, conn, params=(start_date, end_date))
        for col in ["ready_for_dispense_dt", "admin_given_dt", "prepared_dt", "required_start_dt", "event_date"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")
        if "order_id" in df.columns:
            df["order_id"] = df["order_id"].apply(normalize_identifier_text)
        dedupe_cols = [
            c for c in [
                "order_id", "order_medication", "ready_for_dispense_dt", "admin_given_dt",
                "prepared_dt", "prep_or_dispense_user", "pharmacy", "event_date",
                "required_start_dt", "prep_lead_hours", "hold_hours", "is_sjs_cleanroom"
            ] if c in df.columns
        ]
        if dedupe_cols:
            df = df.drop_duplicates(subset=dedupe_cols, keep="first").copy()
        return df


@st.cache_data(ttl=300)
def get_cartfill_available_range():
    init_db()
    query = """
        SELECT MIN(ready_for_dispense_dt)::date AS min_date,
               MAX(ready_for_dispense_dt)::date AS max_date,
               COUNT(*) AS row_count
        FROM wcc_cartfill_stats
        WHERE ready_for_dispense_dt IS NOT NULL
    """
    try:
        with db_cursor() as (conn, cur):
            cur.execute(query)
            row = cur.fetchone()
        if row and row[0] and row[1]:
            return row[0], row[1], row[2] or 0
    except Exception:
        return None, None, 0
    return None, None, 0


@st.cache_data(ttl=3600)
def load_overnight_cartfill_context():
    tables = {
        "windows": "SELECT pk, cartfill_name, time_processed_raw, doses_due, pharmacy FROM overnight_iv_cartfill_windows",
        "staffing": """SELECT pk, schedule_date, day_name, shift_name, weekend_shift_label,
                              assigned_staff, is_weekend, is_placeholder
                       FROM overnight_iv_staffing_model""",
    }
    results = {}
    with db_cursor() as (conn, cur):
        for key, sql in tables.items():
            try:
                results[key] = pd.read_sql(sql, conn)
            except Exception:
                results[key] = pd.DataFrame()
    if "schedule_date" in results.get("staffing", pd.DataFrame()).columns:
        results["staffing"]["schedule_date"] = pd.to_datetime(results["staffing"]["schedule_date"], errors="coerce")
    return results.get("windows", pd.DataFrame()), results.get("staffing", pd.DataFrame())


def clear_app_upload_caches():
    cached_loaders = [
        load_admin_users,
        load_shift_audit_profiles,
        load_shift_schedule_for_date,
        load_day_events_for_shift_audit,
        load_day_pharmacy_for_shift_audit,
        load_shift_audit_results,
        load_data,
        load_pharmacy_workflow_orders,
        load_pyxis_workflow_events,
        load_iv_room_data,
        load_wcc_compounding_stats,
        load_wcc_cartfill_stats,
        load_overnight_cartfill_orders,
        get_cartfill_available_range,
        load_overnight_cartfill_context,
        get_stats_range,
        get_present_dates,
    ]
    for loader in cached_loaders:
        clear_func = getattr(loader, "clear", None)
        if callable(clear_func):
            clear_func()


@st.cache_data(ttl=300)
def get_stats_range():
    try:
        with db_cursor() as (conn, cur):
            sql = """
                WITH
                event_stats AS (
                    SELECT COUNT(*) AS row_count, MIN(dt::date) AS min_dt, MAX(dt::date) AS max_dt
                    FROM events
                ),
                pharmacy_stats AS (
                    SELECT COUNT(*) AS row_count, MIN(dt::date) AS min_dt, MAX(dt::date) AS max_dt
                    FROM pharmacy_orders
                ),
                schedule_stats AS (
                    SELECT COUNT(*) AS row_count, MIN(dt) AS min_dt, MAX(dt) AS max_dt
                    FROM staff_schedule
                ),
                attendance_stats AS (
                    SELECT COUNT(*) AS row_count, MIN(dt_date) AS min_dt, MAX(dt_date) AS max_dt
                    FROM attendance_punches
                )
                SELECT
                    event_stats.row_count,
                    pharmacy_stats.row_count,
                    schedule_stats.row_count,
                    attendance_stats.row_count,
                    (
                        SELECT MIN(d)
                        FROM (VALUES
                            (event_stats.min_dt),
                            (pharmacy_stats.min_dt),
                            (schedule_stats.min_dt),
                            (attendance_stats.min_dt)
                        ) AS dates(d)
                        WHERE d IS NOT NULL
                    ) AS min_dt,
                    (
                        SELECT MAX(d)
                        FROM (VALUES
                            (event_stats.max_dt),
                            (pharmacy_stats.max_dt),
                            (schedule_stats.max_dt),
                            (attendance_stats.max_dt)
                        ) AS dates(d)
                        WHERE d IS NOT NULL
                    ) AS max_dt
                FROM event_stats, pharmacy_stats, schedule_stats, attendance_stats
            """
            cur.execute(sql)
            row = cur.fetchone()
        if row and row[4] and row[5]:
            return (row[0] or 0), (row[1] or 0), (row[2] or 0), (row[3] or 0), row[4], row[5]
    except Exception:
        return 0, 0, 0, 0, date.today(), date.today()
    return 0, 0, 0, 0, date.today(), date.today()

@st.cache_data(ttl=300)
def get_present_dates(min_dt, max_dt):
    if not min_dt or not max_dt:
        return set()
    sql = """
        SELECT DISTINCT d
        FROM (
            SELECT dt::date AS d FROM events WHERE dt IS NOT NULL
            UNION
            SELECT dt::date AS d FROM pharmacy_orders WHERE dt IS NOT NULL
        ) present
        WHERE d BETWEEN %(start)s AND %(end)s
    """
    df = run_query(sql, params={"start": min_dt, "end": max_dt})
    if not df.empty:
        col_name = df.columns[0]
        df[col_name] = pd.to_datetime(df[col_name], errors='coerce')
        return set(df[col_name].dt.date.dropna())
    return set()


def apply_global_styles():
    st.markdown("""
        <style>
        [data-testid="stSidebarNav"] { display: none; }
        .block-container { padding-top: 1.35rem; padding-bottom: 2rem; }
        [data-testid="stAppViewContainer"] {
            background:
                radial-gradient(circle at top left, rgba(56, 189, 248, 0.1), transparent 22%),
                radial-gradient(circle at top right, rgba(16, 185, 129, 0.08), transparent 20%),
                linear-gradient(180deg, #0b1220 0%, #111827 48%, #162131 100%);
        }
        [data-testid="stSidebar"] {
            background:
                radial-gradient(circle at top right, rgba(34, 197, 94, 0.12), transparent 24%),
                linear-gradient(180deg, #0f172a 0%, #111827 45%, #17212f 100%);
            border-right: 1px solid rgba(148, 163, 184, 0.16);
        }
        [data-testid="stSidebar"] * {
            color: #e5eef8;
        }
        [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p,
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] .stRadio label,
        [data-testid="stSidebar"] .stSelectbox label,
        [data-testid="stSidebar"] .stDateInput label,
        [data-testid="stSidebar"] .stSlider label,
        [data-testid="stSidebar"] .stFileUploader label,
        [data-testid="stSidebar"] .stCaptionContainer,
        [data-testid="stSidebar"] small {
            color: #dbe7f3 !important;
        }
        [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] a,
        [data-testid="stSidebar"] [data-testid="stPageLink-NavLink"] {
            color: #f8fafc !important;
        }
        [data-testid="stSidebar"] [data-testid="stPageLink-NavLink"] {
            border-radius: 12px;
            padding: 0.2rem 0.35rem;
        }
        [data-testid="stSidebar"] [data-testid="stPageLink-NavLink"]:hover {
            background: rgba(148, 163, 184, 0.12);
        }
        [data-testid="stSidebar"] .stSelectbox > div > div,
        [data-testid="stSidebar"] .stDateInput > div > div,
        [data-testid="stSidebar"] .stFileUploader > div,
        [data-testid="stSidebar"] .stMultiSelect > div > div {
            background: rgba(15, 23, 42, 0.72);
            border-color: rgba(148, 163, 184, 0.28);
            color: #f8fafc;
        }
        .rx-shell {
            background: linear-gradient(135deg, #0f172a 0%, #1f2937 100%);
            color: white;
            padding: 16px 18px;
            border-radius: 16px;
            margin-bottom: 14px;
            box-shadow: 0 12px 28px rgba(15, 23, 42, 0.16);
        }
        .rx-shell h2 {
            margin: 0 0 4px 0;
            font-size: 1.25rem;
            font-weight: 800;
        }
        .rx-shell p {
            margin: 0;
            color: rgba(255,255,255,0.78);
            font-size: 0.88rem;
            line-height: 1.35;
        }
        .rx-nav-label {
            font-size: 0.72rem;
            font-weight: 800;
            text-transform: uppercase;
            letter-spacing: 0.08em;
            color: #93c5fd;
            margin: 14px 0 6px 0;
        }
        .rx-page-hero {
            background:
                radial-gradient(circle at top right, rgba(59, 130, 246, 0.2), transparent 24%),
                linear-gradient(135deg, #ffffff 0%, #f8fbff 48%, #eef6ff 100%);
            border: 1px solid rgba(148, 163, 184, 0.2);
            border-radius: 22px;
            padding: 1.3rem 1.4rem;
            margin: 0.15rem 0 1.1rem 0;
            box-shadow: 0 18px 40px rgba(15, 23, 42, 0.08);
        }
        .rx-page-kicker {
            display: inline-flex;
            align-items: center;
            gap: 0.35rem;
            font-size: 0.74rem;
            font-weight: 800;
            text-transform: uppercase;
            letter-spacing: 0.08em;
            color: #2563eb;
            background: rgba(37, 99, 235, 0.08);
            border: 1px solid rgba(37, 99, 235, 0.12);
            border-radius: 999px;
            padding: 0.34rem 0.6rem;
            margin-bottom: 0.85rem;
        }
        .rx-page-title {
            margin: 0;
            color: #0f172a;
            font-size: clamp(1.8rem, 3vw, 2.5rem);
            line-height: 1.05;
            font-weight: 900;
            letter-spacing: -0.03em;
        }
        .rx-page-subtitle {
            margin: 0.65rem 0 0 0;
            color: #475569;
            max-width: 72ch;
            font-size: 1rem;
            line-height: 1.6;
        }
        .rx-page-hero .rx-page-title {
            color: #0f172a !important;
        }
        .rx-page-hero .rx-page-subtitle,
        .rx-page-hero p {
            color: #334155 !important;
        }
        [data-testid="stTabs"] [role="tablist"] {
            gap: 0.45rem;
        }
        [data-testid="stTabs"] [role="tab"] {
            background: #eef4fb;
            border: 1px solid rgba(148, 163, 184, 0.42);
            border-radius: 999px;
            color: #1e293b !important;
            font-weight: 700;
            padding: 0.4rem 0.9rem;
        }
        [data-testid="stTabs"] [role="tab"] p,
        [data-testid="stTabs"] [role="tab"] span,
        [data-testid="stTabs"] [role="tab"] div {
            color: inherit !important;
        }
        [data-testid="stTabs"] [role="tab"]:hover {
            background: #dbeafe;
            color: #0f172a;
        }
        [data-testid="stTabs"] [role="tab"][aria-selected="true"] {
            background: linear-gradient(135deg, #0f766e 0%, #0f5f93 100%);
            border-color: rgba(15, 118, 110, 0.6);
            color: #ffffff;
            box-shadow: 0 8px 18px rgba(15, 23, 42, 0.12);
        }
        [data-testid="stTabs"] [role="tabpanel"] {
            background: linear-gradient(180deg, #ffffff 0%, #f8fbff 100%);
            border: 1px solid rgba(148, 163, 184, 0.2);
            border-radius: 18px;
            padding: 1rem 1rem 0.4rem 1rem;
            margin-top: 0.75rem;
            box-shadow: 0 12px 28px rgba(15, 23, 42, 0.06);
        }
        [data-testid="stTabs"] [role="tabpanel"] h1,
        [data-testid="stTabs"] [role="tabpanel"] h2,
        [data-testid="stTabs"] [role="tabpanel"] h3,
        [data-testid="stTabs"] [role="tabpanel"] p,
        [data-testid="stTabs"] [role="tabpanel"] label,
        [data-testid="stTabs"] [role="tabpanel"] span {
            color: #0f172a !important;
        }
        [data-testid="stTabs"] [role="tabpanel"] .stCaption,
        [data-testid="stTabs"] [role="tabpanel"] [data-testid="stCaptionContainer"] {
            color: #475569 !important;
        }
        [data-testid="stTabs"] [role="tabpanel"] .stSelectbox label,
        [data-testid="stTabs"] [role="tabpanel"] .stMultiSelect label,
        [data-testid="stTabs"] [role="tabpanel"] .stNumberInput label,
        [data-testid="stTabs"] [role="tabpanel"] .stDateInput label {
            color: #0f172a !important;
            font-weight: 700;
        }
        [data-testid="stExpander"] details {
            background: linear-gradient(180deg, #ffffff 0%, #f8fbff 100%);
            border: 1px solid rgba(148, 163, 184, 0.2);
            border-radius: 16px;
        }
        [data-testid="stExpander"] summary,
        [data-testid="stExpander"] summary * {
            color: #0f172a !important;
            font-weight: 700;
        }
        [data-testid="stExpander"] [data-testid="stMarkdownContainer"] p,
        [data-testid="stExpander"] [data-testid="stMarkdownContainer"] li,
        [data-testid="stExpander"] [data-testid="stMarkdownContainer"] label,
        [data-testid="stExpander"] [data-testid="stMarkdownContainer"] span,
        [data-testid="stExpander"] .stCaption,
        [data-testid="stExpander"] [data-testid="stCaptionContainer"] {
            color: #334155 !important;
        }
        [data-testid="stExpander"] h1,
        [data-testid="stExpander"] h2,
        [data-testid="stExpander"] h3,
        [data-testid="stExpander"] strong {
            color: #0f172a !important;
        }
        h2, h3, [data-testid="stMarkdownContainer"] h2, [data-testid="stMarkdownContainer"] h3 {
            color: #f8fafc;
        }
        .stCaption, [data-testid="stCaptionContainer"] {
            color: #cbd5e1;
        }
        [data-testid="stMarkdownContainer"] p,
        [data-testid="stMarkdownContainer"] li,
        [data-testid="stMarkdownContainer"] label {
            color: #e2e8f0;
        }
        </style>
    """, unsafe_allow_html=True)


def get_management_password():
    """Return the configured management password without exposing it in the UI."""
    env_password = os.environ.get("RXTRACK_MANAGEMENT_PASSWORD")
    if env_password:
        return str(env_password)

    secret_paths = [
        ("management", "password"),
        ("rxtrack", "management_password"),
    ]
    for section, key in secret_paths:
        try:
            value = st.secrets.get(section, {}).get(key)
            if value:
                return str(value)
        except Exception:
            pass

    try:
        value = st.secrets.get("management_password")
        if value:
            return str(value)
    except Exception:
        pass
    return "1234"


def management_access_unlocked():
    return bool(st.session_state.get("_rxtrack_management_unlocked", False))


def render_management_login(page_name="Management"):
    password = get_management_password()
    if management_access_unlocked():
        return True

    st.warning(f"{page_name} is password protected.")
    with st.form(f"management_login_{page_name.replace(' ', '_').lower()}"):
        entered = st.text_input("Management password", type="password")
        submitted = st.form_submit_button("Unlock")
    if submitted:
        if hmac.compare_digest(entered or "", password):
            st.session_state["_rxtrack_management_unlocked"] = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    return False


def require_management_access(page_name="Management"):
    if not render_management_login(page_name):
        st.stop()


def render_management_logout():
    if management_access_unlocked():
        if st.button("Lock Management", key="rxtrack_management_logout"):
            st.session_state["_rxtrack_management_unlocked"] = False
            st.rerun()


def render_management_sidebar_gate():
    if management_access_unlocked():
        st.caption("Management unlocked")
        return
    with st.expander("Management", expanded=False):
        render_management_login("Management")


def render_page_links():
    def safe_page_link(path, label, icon):
        if path == "App.py" or os.path.exists(path):
            st.page_link(path, label=label, icon=icon)

    st.markdown('<div class="rx-nav-label">Core</div>', unsafe_allow_html=True)
    safe_page_link("App.py", label="Overview Hub", icon="🏠")
    safe_page_link("pages/Projects_Portfolio.py", label="Projects Portfolio", icon="📁")
    safe_page_link("pages/🧪_Workflow_Experiments.py", label="Workflow Experiments", icon="🧪")
    safe_page_link("pages/🧭_Pilot_Monitor.py", label="Pilot Monitor", icon="🧭")
    safe_page_link("pages/⚖️_Workload_Capacity_Simulator.py", label="Ops Simulator", icon="⚖️")

    st.markdown('<div class="rx-nav-label">Operations</div>', unsafe_allow_html=True)
    safe_page_link("pages/🏥_Pharmacy_Workflow.py", label="Pharmacy Workflow", icon="🏥")
    safe_page_link("pages/💉_IV_Room.py", label="IV Room", icon="💉")
    safe_page_link("pages/WCC.py", label="WCC", icon="🍼")
    safe_page_link("pages/Device_Utilization.py", label="Device Utilization", icon="📟")
    safe_page_link("pages/🌙_Cartfill_Optimizer.py", label="Cartfill Optimizer", icon="🌙")
    safe_page_link("pages/🔄_Return_Reconciliation.py", label="Return Reconciliation", icon="🔄")
    safe_page_link("pages/🗑️_Return_Bin_Tracker.py", label="Return Bin & Cassettes", icon="🗑️")
    safe_page_link("pages/Inventory_Quality_Control.py", label="Inventory Quality Control", icon="📦")
    safe_page_link("pages/Mobile_BUD_Scanner.py", label="Mobile BUD Scanner", icon="📱")

    if management_access_unlocked():
        st.markdown('<div class="rx-nav-label">Management</div>', unsafe_allow_html=True)
        safe_page_link("pages/🎯_Daily_Command.py", label="Daily Command", icon="🎯")
        safe_page_link("pages/1_⏰_Tardies.py", label="Tardies", icon="⏰")
        safe_page_link("pages/Management_Coaching.py", label="Coaching", icon="📝")
        st.markdown('<div class="rx-nav-label">Performance</div>', unsafe_allow_html=True)
        safe_page_link("pages/2_🔍_Session_Explorer.py", label="Session Explorer", icon="🔍")
        safe_page_link("pages/📈_Shift_Audit_Monitor.py", label="Shift Audit Monitor", icon="📈")
        safe_page_link("pages/📊_Workforce_Intelligence.py", label="Workforce Intelligence", icon="📊")
        safe_page_link("pages/📥_Pends_Analyzer.py", label="Pends Analyzer", icon="📥")
        safe_page_link("pages/🚨_discrepancy_deep_dive.py", label="Discrepancy Deep Dive", icon="🚨")

    st.markdown('<div class="rx-nav-label">Tools</div>', unsafe_allow_html=True)
    safe_page_link("pages/📊_Cycle_Count_Integrity.py", label="Cycle Count Integrity", icon="📊")
    safe_page_link("pages/📋_Carousel_Drop_Tracker.py", label="Carousel Drop Tracker", icon="📋")
    safe_page_link("pages/_🔍_MedLookup.py", label="Med Lookup", icon="🔍")
    safe_page_link("pages/🗄️_db_health.py", label="Database Health", icon="🗄️")
    safe_page_link("pages/Admin_Master_Mapping.py", label="Admin & Mapping", icon="⚙️")


def render_sidebar_chrome():
    """Use this on pages that need the shared nav styling without the date filters."""
    init_db()
    apply_global_styles()
    with st.sidebar:
        st.markdown("""
            <div class="rx-shell">
                <h2>RxTrack</h2>
                <p>Operations, analytics, staffing pilots, and workflow testing in one place.</p>
            </div>
        """, unsafe_allow_html=True)
        render_page_links()
        st.divider()
        render_management_sidebar_gate()
        render_management_logout()
        render_demo_mode_toggle()
        render_ui_debug_toggle()


def render_page_intro(title, subtitle=None, kicker="Operations Intelligence"):
    """Shared page header so every section matches the new RxTrack shell."""
    apply_global_styles()
    subtitle_html = f'<p class="rx-page-subtitle">{subtitle}</p>' if subtitle else ""
    st.markdown(
        f"""
        <section class="rx-page-hero">
            <div class="rx-page-kicker">{kicker}</div>
            <h1 class="rx-page-title">{title}</h1>
            {subtitle_html}
        </section>
        """,
        unsafe_allow_html=True,
    )


def ui_debug_enabled():
    query_enabled = False
    try:
        qp = st.query_params
        query_enabled = str(qp.get("ui_debug", "0")).lower() in {"1", "true", "yes", "on"}
    except Exception:
        query_enabled = False
    if query_enabled:
        st.session_state["_ui_debug_enabled"] = True
    return bool(st.session_state.get("_ui_debug_enabled", False))


def render_ui_debug_toggle():
    current = ui_debug_enabled()
    enabled = st.toggle("UI Debug Mode", value=current, key="rxtrack_ui_debug_toggle")
    st.session_state["_ui_debug_enabled"] = enabled


def demo_mode_enabled():
    return bool(st.session_state.get("_rxtrack_demo_mode", False))


def render_demo_mode_toggle():
    current = demo_mode_enabled()
    enabled = st.toggle("Interview Demo Mode", value=current, key="rxtrack_demo_mode_toggle")
    st.session_state["_rxtrack_demo_mode"] = enabled
    if enabled:
        st.caption("Demo mode highlights the strongest workflows and quiets the internal admin feel.")


def render_demo_hub(start_date, end_date, df_events, df_pharm, df_sched, df_att):
    st.markdown("## Interview Demo")
    st.caption(
        f"Story window: {start_date.strftime('%b %d, %Y')} to {end_date.strftime('%b %d, %Y')} | "
        "Use this as a guided walkthrough of workflow intelligence, staffing analytics, and operational waste reduction."
    )

    ev_count = len(df_events) if not df_events.empty else 0
    pharm_count = len(df_pharm) if not df_pharm.empty else 0
    sched_count = len(df_sched) if not df_sched.empty else 0
    att_count = len(df_att) if not df_att.empty else 0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Pyxis Events", f"{ev_count:,}")
    m2.metric("Pharmacy Orders", f"{pharm_count:,}")
    m3.metric("Schedule Rows", f"{sched_count:,}")
    m4.metric("Attendance Punches", f"{att_count:,}")

    st.divider()

    st.subheader("Recommended Walkthrough")
    walkthrough = [
        (
            "1. Return Reconciliation",
            "Show chain-of-custody logic from Pyxis removals to carousel returns.",
            "pages/🔄_Return_Reconciliation.py",
            "Open Return Reconciliation",
            "🔄",
        ),
        (
            "2. Pends Analyzer",
            "Highlight reload churn, boomerang meds, machine-specific waste, and raw verification drilldowns.",
            "pages/📥_Pends_Analyzer.py",
            "Open Pends Analyzer",
            "📥",
        ),
        (
            "3. IV Room",
            "Walk through daily workload, drill into a production day, and inspect unassigned prep rows.",
            "pages/💉_IV_Room.py",
            "Open IV Room",
            "💉",
        ),
        (
            "4. Tardies / Schedule Logic",
            "Show shift-trade handling, color-aware schedule exceptions, and attendance analytics.",
            "pages/1_⏰_Tardies.py",
            "Open Tardies",
            "⏰",
        ),
    ]

    for title, desc, path, label, icon in walkthrough:
        c1, c2 = st.columns([3, 1])
        with c1:
            st.markdown(f"**{title}**")
            st.caption(desc)
        with c2:
            st.page_link(path, label=label, icon=icon)

    st.divider()

    c1, c2 = st.columns(2)
    with c1:
        with st.expander("3-Minute Interview Script", expanded=True):
            st.markdown(
                """
1. RxTrack unifies Pyxis, carousel, IV room, scheduling, and attendance data into one operational app.
2. It was built to surface waste, missed reconciliation, staffing exceptions, and repeat manual work.
3. Return Reconciliation shows whether removals from Pyxis are actually matched by carousel-side return workflows.
4. Pends Analyzer shows medications that get unloaded and then reloaded back into the same machine, which points to standard-stock opportunities.
5. IV Room turns raw preparation logs into staffing and production visibility with day-level drilldowns.
                """
            )
    with c2:
        with st.expander("10-Minute Demo Narrative", expanded=True):
            st.markdown(
                """
- Start with the Overview Hub to frame the app as pharmacy operations intelligence.
- Go to Return Reconciliation and explain closed-loop medication accountability.
- Move to Pends Analyzer and show boomerang meds by machine, raw supporting events, and quick-correction filtering.
- Open IV Room and drill into one day to show production analytics from raw event data.
- Finish with Tardies to show schedule-aware attendance logic, including trade and adjustment handling.
                """
            )

    with st.expander("What This Demonstrates", expanded=False):
        st.markdown(
            """
- Workflow understanding, not just dashboarding
- Translating frontline pain points into operational tooling
- Iterating logic when real-world pharmacy processes do not fit a naive model
- Building analytics that support both accountability and efficiency
            """
        )


def record_ui_debug_event(page_name, event, **details):
    if not ui_debug_enabled():
        return
    history = st.session_state.setdefault("_ui_debug_history", [])
    history.append(
        {
            "page": page_name,
            "event": event,
            "details": details,
        }
    )
    st.session_state["_ui_debug_history"] = history[-25:]


def render_ui_debugger(page_name, intro_mode=None, extra=None):
    if not ui_debug_enabled():
        return
    history = st.session_state.get("_ui_debug_history", [])
    payload = {
        "page": page_name,
        "intro_mode": intro_mode,
        "has_render_sidebar_chrome": "render_sidebar_chrome" in globals(),
        "has_render_page_intro": "render_page_intro" in globals(),
        "history_size": len(history),
    }
    if extra:
        payload.update(extra)
    with st.sidebar.expander("UI Debugger", expanded=True):
        st.code(json.dumps(payload, indent=2), language="json")
        if history:
            st.caption("Recent UI events")
            for item in reversed(history[-8:]):
                st.write(f"- {item['page']}: {item['event']} {item['details']}")


# --- SHARED SIDEBAR RENDERER ---
def render_sidebar():
    """Call this at the top of any page to always show the date range sidebar."""
    init_db()
    apply_global_styles()
    n_events, n_pharm, n_sched, n_att, min_db, max_db = get_stats_range()
    today = date.today()
    min_selectable_date = min(min_db, today)
    max_selectable_date = max(max_db, today)

    if 'start_date' not in st.session_state:
        st.session_state.start_date = today
    if 'end_date' not in st.session_state:
        st.session_state.end_date = today
    if 'rxtrack_sidebar_filter_mode' not in st.session_state:
        st.session_state.rxtrack_sidebar_filter_mode = "Day"
    st.session_state.start_date = min(max(st.session_state.start_date, min_selectable_date), max_selectable_date)
    st.session_state.end_date = min(max(st.session_state.end_date, min_selectable_date), max_selectable_date)
    if st.session_state.start_date > st.session_state.end_date:
        st.session_state.end_date = st.session_state.start_date

    with st.sidebar:
        st.markdown("""
            <div class="rx-shell">
                <h2>RxTrack</h2>
                <p>Operations, analytics, staffing pilots, and workflow testing in one place.</p>
            </div>
        """, unsafe_allow_html=True)
        render_page_links()

        st.divider()
        render_management_sidebar_gate()
        render_management_logout()
        render_demo_mode_toggle()
        render_ui_debug_toggle()
        st.markdown("### Analysis Window")

        filter_mode = st.radio(
            "Filter Mode",
            ["Range", "Month", "Week", "Day"],
            horizontal=True,
            label_visibility="collapsed",
            key="rxtrack_sidebar_filter_mode",
        )

        if filter_mode == "Range":
            range_start, range_end = st.columns(2)
            selected_start = range_start.date_input(
                "Start Date",
                value=st.session_state.start_date,
                min_value=min_selectable_date,
                max_value=max_selectable_date,
                key="rxtrack_sidebar_range_start",
            )
            selected_end = range_end.date_input(
                "End Date",
                value=st.session_state.end_date,
                min_value=min_selectable_date,
                max_value=max_selectable_date,
                key="rxtrack_sidebar_range_end",
            )
            if selected_start > selected_end:
                st.warning("Start date was after end date, so RxTrack used the same day for both.")
                selected_end = selected_start
            st.session_state.start_date = selected_start
            st.session_state.end_date = selected_end

        elif filter_mode == "Month":
            selected_month = st.date_input(
                "Select Month:",
                value=st.session_state.start_date,
                min_value=min_selectable_date,
                max_value=max_selectable_date,
            )
            month_start = selected_month.replace(day=1)
            next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
            st.session_state.start_date = month_start
            st.session_state.end_date = next_month - timedelta(days=1)

        elif filter_mode == "Week":
            week_start = st.date_input(
                "Select Week Start:",
                value=st.session_state.start_date,
                min_value=min_selectable_date,
                max_value=max_selectable_date,
            )
            st.session_state.start_date = week_start
            st.session_state.end_date = week_start + timedelta(days=6)

        else:
            single_day = st.date_input(
                "Select Day:",
                value=st.session_state.start_date,
                min_value=min_selectable_date,
                max_value=max_selectable_date,
            )
            st.session_state.start_date = single_day
            st.session_state.end_date = single_day

        st.divider()

        with st.expander("Database Status", expanded=False):
            c1, c2 = st.columns(2)
            c1.metric("Pyxis Events", f"{n_events:,}")
            c2.metric("Pharm Orders", f"{n_pharm:,}")
            c3, c4 = st.columns(2)
            c3.metric("Sched. Shifts", f"{n_sched:,}")
            c4.metric("Time Punches", f"{n_att:,}")

            if min_db and max_db and min_db <= max_db:
                delta = (max_db - min_db).days
                cal_start = max_db - timedelta(days=90) if delta > 90 else min_db
                present_dates = get_present_dates(cal_start, max_db)
                cal_html = '<div class="cal-grid">'
                curr = cal_start
                while curr <= max_db:
                    color = "cal-present" if curr in present_dates else "cal-missing"
                    cal_html += f'<div class="cal-day {color}" title="{curr}"></div>'
                    curr += timedelta(days=1)
                cal_html += '</div>'
                st.markdown(cal_html, unsafe_allow_html=True)

    return st.session_state.start_date, st.session_state.end_date



# --- MAIN APP LOGIC ---
# Only runs when App.py is the active Streamlit entrypoint.
# When other pages import App.py, they should only get the shared helpers above.
_is_main = (__name__ == "__main__")

if _is_main:
    # Apply page config and CSS only when running as main page

    # --- CONFIGURATION ---
    st.set_page_config(
        page_title="RxTrack: Workforce & Efficiency", 
        page_icon="🏥",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    apply_global_styles()

    # Suppress DB/Pandas warnings
    warnings.filterwarnings("ignore", category=UserWarning, module="pandas")

    # --- INITIALIZE VARIABLES (Prevents NameError) ---
    df_events = pd.DataFrame()
    df_config = pd.DataFrame()
    df_pharm = pd.DataFrame()
    df_sched = pd.DataFrame()
    df_att = pd.DataFrame()

    # --- CONSTANTS ---
    NARC_TERMS = [
        "OXYCODONE", "MORPHINE", "FENTANYL", "HYDROMORPHONE", "HYDROCODONE", 
        "LORAZEPAM", "MIDAZOLAM", "DIAZEPAM", "ALPRAZOLAM", "CODEINE", 
        "METHADONE", "KETAMINE", "TRAMADOL", "ZOLPIDEM", "PHENOBARBITAL", 
        "BUPRENORPHINE", "LACOSAMIDE", "VIMPAT", "PREGABALIN", "LYRICA", 
        "CHLORDIAZEPOXIDE", "LIBRIUM", "CLONAZEPAM", "KLONOPIN"
    ]

    ADMIN_USERS = load_admin_users()

    # --- CUSTOM CSS ---
    st.markdown("""
        <style>
        .metric-card { 
            background-color: #ffffff; 
            padding: 20px; 
            border-radius: 10px; 
            border-left: 5px solid #4CAF50; 
            box-shadow: 0 4px 6px rgba(0,0,0,0.1); 
            margin-bottom: 10px;
        }
        .metric-card h3 { 
            color: #1f2937; 
            margin: 0; 
            font-size: 26px; 
            font-weight: 700; 
        }
        .metric-card p { 
            color: #6b7280; 
            margin: 0; 
            font-size: 14px; 
            font-weight: 500; 
            text-transform: uppercase; 
            letter-spacing: 0.5px;
        }
        .cal-grid { display: flex; flex-wrap: wrap; gap: 3px; max-width: 100%; margin-top: 10px; }
        .cal-day { 
            width: 12px; height: 12px; 
            border-radius: 2px; 
            background-color: #e5e7eb;
        }
        .cal-present { background-color: #4CAF50; } /* Green */
        .cal-missing { background-color: #F87171; } /* Red */
        </style>
        """, unsafe_allow_html=True)

    # --- MAIN APP LOGIC ---
    init_db()

    # 1. Define your internal pages (those not yet moved to the /pages folder)
    PAGES = [
        "📊 Overview", "🛡️ Compliance", "🚚 Load/Unload",
    ]

    # Use render_sidebar for date range — same as all other pages
    start_date, end_date = render_sidebar()

    with st.sidebar:
        st.markdown("### 🧭 Navigation")
        selected_page = st.radio("Go to:", PAGES, label_visibility="collapsed")
        st.divider()

        # --- UNIVERSAL DATA INGEST ---
        st.subheader("📤 Ingest Data")
        u_type = st.selectbox("File Type:", [
            "Daily Transaction Report", "Device Activity Log (Pends)", "Pharmacy Workflow Report", 
            "Med Cost Prices (Legacy Inventory Audit)", "Detailed Inventory Snapshot (Legacy RC)", "Staff Schedule", "Attendance Tracking",
            "IV Room Workload", "IV Room Batching", "IV Room Workflow Detail",
            "WCC Compounding Stats", "Cartfill Stats (All Areas)", "WCC Cartfill Stats",
            "Days Since Last Cycle Count Report", "Cycle Count Variance Report",
            "Buyer Formulary Listing Report", "Physical Inventory Report",
            "Audit Transaction Detail RC",
            "Packaging Report", "Device Inventory List"
        ])
        if u_type == "Audit Transaction Detail RC":
            st.caption("This upload also refreshes med cost prices from the UnitCost column.")
        elif u_type == "Med Cost Prices (Legacy Inventory Audit)":
            st.caption("Legacy price-only upload. Audit Transaction Detail RC now refreshes med costs too.")
        elif u_type == "Detailed Inventory Snapshot (Legacy RC)":
            st.caption("Legacy current-inventory snapshot. Device Inventory List is the preferred current Pyxis inventory source.")
        upload_types = None if u_type == "Packaging Report" else ["csv", "xlsx"]
        uploaded_files = st.file_uploader(f"Upload {u_type}", type=upload_types, accept_multiple_files=True)

        class LocalReportFile:
            def __init__(self, path):
                self.path = path
                self.name = os.path.basename(path)
                self._handle = open(path, "rb")

            def read(self, *args, **kwargs):
                return self._handle.read(*args, **kwargs)

            def readline(self, *args, **kwargs):
                return self._handle.readline(*args, **kwargs)

            def seek(self, *args, **kwargs):
                return self._handle.seek(*args, **kwargs)

            def close(self):
                return self._handle.close()

        report_folder = st.text_input(
            "Optional report folder scan (UNC path)",
            value=r"\\ILXFSCIFS01.il.hshsad.org\PyxisES$\Reports\JMWolfe",
            help="Tries this hospital network path first. This only works if RxTrack is running locally on a work-network PC that can access the share.",
        )
        mapped_report_folder = st.text_input(
            "Mapped drive path",
            value=r"Z:\Reports\JMWolfe",
            help="Optional fallback. If the UNC path fails, map the network share to a drive letter and enter that path here.",
        )
        folder_filename_filter = st.text_input(
            "Folder filename contains",
            value="",
            help="Optional. Use this when the folder contains several report types, for example `Workflow` or `Transaction`.",
        )

        def folder_access_help_message(primary_path, fallback_path=None):
            fallback_text = f"\n- Mapped drive fallback tried: `{fallback_path}`" if fallback_path else ""
            return (
                "RxTrack could not access the report folder from the machine running the app.\n\n"
                f"- UNC path tried: `{primary_path}`"
                f"{fallback_text}\n"
                "- The app must be running locally on your hospital work PC or another machine inside the hospital network.\n"
                "- Streamlit Cloud cannot access hospital network shares like `\\\\ILXFSCIFS01...`.\n"
                "- If the UNC path fails locally, map the network folder to a drive letter such as `Z:` and use the mapped drive path."
            )

        def list_report_folder_files(folder_path, name_filter=""):
            if not os.path.isdir(folder_path):
                raise FileNotFoundError(f"Path does not exist or is not accessible: {folder_path}")
            allowed_exts = {".csv", ".xlsx", ".xls", ".txt"}
            name_filter = str(name_filter or "").strip().lower()
            paths = []
            try:
                names = os.listdir(folder_path)
            except Exception as list_error:
                raise PermissionError(f"Path exists but RxTrack could not list files: {folder_path}. {list_error}") from list_error
            for name in names:
                path = os.path.join(folder_path, name)
                if name_filter and name_filter not in name.lower():
                    continue
                if os.path.isfile(path) and os.path.splitext(name)[1].lower() in allowed_exts:
                    paths.append(path)
            return sorted(paths, key=lambda path: os.path.getmtime(path))

        def find_accessible_report_files(primary_path, fallback_path="", name_filter=""):
            attempts = []
            for label, path in [("UNC path", primary_path), ("Mapped drive path", fallback_path)]:
                clean_path = str(path or "").strip()
                if not clean_path:
                    continue
                try:
                    return clean_path, list_report_folder_files(clean_path, name_filter), attempts
                except Exception as path_error:
                    attempts.append((label, clean_path, str(path_error)))
            raise FileNotFoundError(folder_access_help_message(primary_path, fallback_path))

        test_folder_clicked = st.button("Test folder access")
        if test_folder_clicked:
            try:
                active_folder, folder_paths, attempts = find_accessible_report_files(
                    report_folder,
                    mapped_report_folder,
                    folder_filename_filter,
                )
                st.success(f"Folder access works: `{active_folder}`")
                if attempts:
                    for label, path, error_text in attempts:
                        st.warning(f"{label} failed for `{path}`. Fallback succeeded. Details: {error_text}")
                if folder_paths:
                    st.info(f"Found {len(folder_paths)} matching CSV/XLSX/TXT file{'s' if len(folder_paths) != 1 else ''}. Showing first 10.")
                    st.dataframe(
                        pd.DataFrame(
                            [
                                {
                                    "file_name": os.path.basename(path),
                                    "modified": datetime.fromtimestamp(os.path.getmtime(path)),
                                    "path": path,
                                }
                                for path in folder_paths[:10]
                            ]
                        ),
                        width="stretch",
                        hide_index=True,
                    )
                else:
                    st.warning("Folder is accessible, but no CSV/XLSX/TXT files matched the current filename filter.")
            except Exception as folder_error:
                st.error(str(folder_error))

        folder_scan_clicked = st.button(f"Scan folder and process {u_type}")
        files_to_process = []
        folder_file_handles = []
        if folder_scan_clicked:
            try:
                active_folder, folder_paths, attempts = find_accessible_report_files(
                    report_folder,
                    mapped_report_folder,
                    folder_filename_filter,
                )
                if folder_paths:
                    folder_file_handles = [LocalReportFile(path) for path in folder_paths]
                    files_to_process = folder_file_handles
                    st.success(f"Using report folder: `{active_folder}`")
                    if attempts:
                        for label, path, error_text in attempts:
                            st.warning(f"{label} failed for `{path}`. Fallback succeeded. Details: {error_text}")
                    st.info(f"Found {len(files_to_process)} report file{'s' if len(files_to_process) != 1 else ''} in the folder.")
                else:
                    st.warning("Folder is accessible, but no CSV/XLSX/TXT files matched the current filename filter.")
            except Exception as folder_error:
                st.error(str(folder_error))

        def read_uploaded_tabular(file):
            file.seek(0)
            if file.name.endswith('.xlsx'):
                return pd.read_excel(file)
            try:
                return pd.read_csv(file, low_memory=False)
            except UnicodeDecodeError:
                file.seek(0)
                return pd.read_csv(file, encoding='latin1', low_memory=False)

        def read_uploaded_batch(files):
            frames = [read_uploaded_tabular(file) for file in files]
            return pd.concat(frames, ignore_index=True) if len(frames) > 1 else frames[0]
        device_inventory_snapshot_date = date.today()
        update_current_device_inventory = True
        if u_type == "Device Inventory List":
            device_inventory_snapshot_date = st.date_input(
                "Device Inventory snapshot date",
                value=date.today(),
                help="Use this to backfill older Device Inventory CSVs so daily movement can compare the correct days.",
            )
            update_current_device_inventory = st.checkbox(
                "Replace current Device Inventory table with this file",
                value=device_inventory_snapshot_date == date.today(),
                help="Leave this unchecked when backfilling older files. The file will still be saved to daily history.",
            )
        process_label = f"Process {len(uploaded_files)} {u_type} file{'s' if len(uploaded_files) != 1 else ''}" if uploaded_files else f"Process {u_type}"
        if uploaded_files and st.button(process_label):
            files_to_process = list(uploaded_files)

        if files_to_process:
            uploaded_files = files_to_process
            uploaded = uploaded_files[0]
            uploaded_names = ", ".join(file.name for file in uploaded_files)
            try:
                upload_started = time.perf_counter()
                processed_count = 0
                # 1. Load raw file
                direct_file_reports = {
                    "Days Since Last Cycle Count Report",
                    "Cycle Count Variance Report",
                    "Buyer Formulary Listing Report",
                    "Physical Inventory Report",
                    "Packaging Report",
                }
                if u_type in direct_file_reports:
                    raw = None
                else:
                    raw = read_uploaded_batch(uploaded_files)

                # 2. Route to correct SQL Table
                clean = None 

                if u_type == "Daily Transaction Report":
                    clean = clean_dataframe(raw)
                    sql = """INSERT INTO events (pk, user_name, device, med_id, med_desc, event_type, dt, qty, 
                             beginning_qty, ending_qty, discrepancy_qty, discrepancy_reason, resolution_dt) 
                             VALUES (%(pk)s, %(user_name)s, %(device)s, %(med_id)s, %(med_desc)s, %(event_type)s, 
                             %(dt)s, %(qty)s, %(beginning_qty)s, %(ending_qty)s, %(discrepancy_qty)s, 
                             %(discrepancy_reason)s, %(resolution_dt)s) ON CONFLICT (pk) DO NOTHING;"""
                    execute_statement(sql, clean.to_dict("records"), batch=True, table_name="Events")

                elif u_type == "Device Activity Log (Pends)":
                    clean = clean_activity_log(raw)
                    sql = """INSERT INTO config_events (pk, dt, user_name, device, med_id, location, 
                             action_type, activity_category, min_qty, max_qty, is_standard) 
                             VALUES (%(pk)s, %(dt)s, %(user_name)s, %(device)s, %(med_id)s, %(location)s, 
                             %(action_type)s, %(activity_category)s, %(min_qty)s, %(max_qty)s, %(is_standard)s) 
                             ON CONFLICT (pk) DO NOTHING;"""
                    execute_statement(sql, clean.to_dict("records"), batch=True, table_name="Config")

                elif u_type == "Pharmacy Workflow Report":
                    clean = clean_pharmacy_report(raw)
                    sql = """INSERT INTO pharmacy_orders (pk, queue_id, priority, dt, med_id, med_desc, 
                             destination, user_name, qty) VALUES (%(pk)s, %(queue_id)s, %(priority)s, 
                             %(dt)s, %(med_id)s, %(med_desc)s, %(destination)s, %(user_name)s, %(qty)s) 
                             ON CONFLICT (pk) DO NOTHING;"""
                    execute_statement(sql, clean.to_dict("records"), batch=True, table_name="Pharmacy Orders")

                elif u_type == "Med Cost Prices (Legacy Inventory Audit)":
                    clean = clean_inventory_file(raw)
                    sql_costs = """INSERT INTO med_costs (med_id, cost_per_unit) VALUES (%(med_id)s, %(unit_cost)s) 
                                   ON CONFLICT (med_id) DO UPDATE SET cost_per_unit = EXCLUDED.cost_per_unit;"""
                    execute_statement(sql_costs, clean.to_dict("records"), batch=True, table_name="Cost Updates")

                elif u_type == "Staff Schedule":
                    if len(uploaded_files) > 1:
                        clean = pd.concat(
                            [
                                clean_schedule_data(read_uploaded_tabular(file), file if file.name.endswith('.xlsx') else None)
                                for file in uploaded_files
                            ],
                            ignore_index=True,
                        )
                    else:
                        clean = clean_schedule_data(raw, uploaded if uploaded.name.endswith('.xlsx') else None)
                    sql = """INSERT INTO staff_schedule (pk, dt, day_name, staff_name, shift_type, 
                             assignment_type, raw_entry, note, schedule_status, cell_fill_color)
                             VALUES (%(pk)s, %(dt)s, %(day_name)s,
                             %(staff_name)s, %(shift_type)s, %(assignment_type)s, %(raw_entry)s, 
                             %(note)s, %(schedule_status)s, %(cell_fill_color)s)
                             ON CONFLICT (pk) DO NOTHING;"""
                    execute_statement(sql, clean.to_dict("records"), batch=True, table_name="Schedule")

                elif u_type == "Attendance Tracking":
                    clean = pd.concat([clean_attendance_file(file) for file in uploaded_files], ignore_index=True)
                    sql = """INSERT INTO attendance_punches (pk, raw_name, dt_date, start_dt, end_dt) 
                             VALUES (%(pk)s, %(raw_name)s, %(dt_date)s, %(start_dt)s, %(end_dt)s) 
                             ON CONFLICT (pk) DO NOTHING;"""
                    execute_statement(sql, clean.to_dict("records"), batch=True, table_name="Attendance")

                elif u_type == "Detailed Inventory Snapshot (Legacy RC)":
                    clean = clean_detailed_inventory(raw)
                    sql = """INSERT INTO inventory_detailed 
                             (pk, station, med_id, med_desc, unit_cost, current_count, pocket_location)
                             VALUES (%(pk)s, %(station)s, %(med_id)s, %(med_desc)s, 
                                     %(unit_cost)s, %(current_count)s, %(pocket_location)s)
                             ON CONFLICT (pk) DO NOTHING;"""
                    execute_statement(sql, clean.to_dict("records"), batch=True, table_name="Detailed Inventory")

                elif u_type == "Device Inventory List":
                    clean = clean_device_inventory(raw)
                    history_sql = """INSERT INTO device_inventory_history
                             (snapshot_date, pk, med_desc, device, zone, pocket_location, status, brand_name, med_id, med_class,
                              current_quantity, min_qty, max_qty, outdate_tracking, loaded_as_fraction,
                              backordered, standard_stock, active_orders, days_unused, snapshot_dt)
                             VALUES (%(snapshot_date)s, %(pk)s, %(med_desc)s, %(device)s, %(zone)s, %(pocket_location)s, %(status)s,
                                     %(brand_name)s, %(med_id)s, %(med_class)s, %(current_quantity)s, %(min_qty)s,
                                     %(max_qty)s, %(outdate_tracking)s, %(loaded_as_fraction)s, %(backordered)s,
                                     %(standard_stock)s, %(active_orders)s, %(days_unused)s, NOW())
                             ON CONFLICT (snapshot_date, pk) DO UPDATE SET
                                 med_desc = EXCLUDED.med_desc,
                                 device = EXCLUDED.device,
                                 zone = EXCLUDED.zone,
                                 pocket_location = EXCLUDED.pocket_location,
                                 status = EXCLUDED.status,
                                 brand_name = EXCLUDED.brand_name,
                                 med_id = EXCLUDED.med_id,
                                 med_class = EXCLUDED.med_class,
                                 current_quantity = EXCLUDED.current_quantity,
                                 min_qty = EXCLUDED.min_qty,
                                 max_qty = EXCLUDED.max_qty,
                                 outdate_tracking = EXCLUDED.outdate_tracking,
                                 loaded_as_fraction = EXCLUDED.loaded_as_fraction,
                                 backordered = EXCLUDED.backordered,
                                 standard_stock = EXCLUDED.standard_stock,
                                 active_orders = EXCLUDED.active_orders,
                                 days_unused = EXCLUDED.days_unused,
                                 snapshot_dt = NOW();"""
                    sql = """INSERT INTO device_inventory
                             (pk, med_desc, device, zone, pocket_location, status, brand_name, med_id, med_class,
                              current_quantity, min_qty, max_qty, outdate_tracking, loaded_as_fraction,
                              backordered, standard_stock, active_orders, days_unused, snapshot_dt)
                             VALUES (%(pk)s, %(med_desc)s, %(device)s, %(zone)s, %(pocket_location)s, %(status)s,
                                     %(brand_name)s, %(med_id)s, %(med_class)s, %(current_quantity)s, %(min_qty)s,
                                     %(max_qty)s, %(outdate_tracking)s, %(loaded_as_fraction)s, %(backordered)s,
                                     %(standard_stock)s, %(active_orders)s, %(days_unused)s, NOW())
                             ON CONFLICT (pk) DO UPDATE SET
                                 med_desc = EXCLUDED.med_desc,
                                 zone = EXCLUDED.zone,
                                 status = EXCLUDED.status,
                                 brand_name = EXCLUDED.brand_name,
                                 med_class = EXCLUDED.med_class,
                                 current_quantity = EXCLUDED.current_quantity,
                                 min_qty = EXCLUDED.min_qty,
                                 max_qty = EXCLUDED.max_qty,
                                 outdate_tracking = EXCLUDED.outdate_tracking,
                                 loaded_as_fraction = EXCLUDED.loaded_as_fraction,
                                 backordered = EXCLUDED.backordered,
                                 standard_stock = EXCLUDED.standard_stock,
                                 active_orders = EXCLUDED.active_orders,
                                 days_unused = EXCLUDED.days_unused,
                                 snapshot_dt = NOW();"""
                    clean_records = clean.to_dict("records")
                    for row in clean_records:
                        row["snapshot_date"] = device_inventory_snapshot_date
                    with db_cursor() as (conn, cur):
                        cur.execute("DELETE FROM device_inventory_history WHERE snapshot_date = %s;", (device_inventory_snapshot_date,))
                        conn.commit()
                    execute_statement(history_sql, clean_records, batch=True, table_name="Device Inventory History")
                    if update_current_device_inventory:
                        with db_cursor() as (conn, cur):
                            cur.execute("DELETE FROM device_inventory;")
                            conn.commit()
                        execute_statement(sql, clean.to_dict("records"), batch=True, table_name="Device Inventory")

                elif u_type == "Days Since Last Cycle Count Report":
                    clean = pd.concat([clean_cycle_count_status_report(file) for file in uploaded_files], ignore_index=True)
                    sql = """INSERT INTO cycle_count_status
                             (pk, snapshot_date, source_filename, isa_name, med_id, med_desc, location,
                              cycle_count_interval, last_cycle_count, days_since_last_count, days_over_due)
                             VALUES (%(pk)s, %(snapshot_date)s, %(source_filename)s, %(isa_name)s, %(med_id)s,
                                     %(med_desc)s, %(location)s, %(cycle_count_interval)s, %(last_cycle_count)s,
                                     %(days_since_last_count)s, %(days_over_due)s)
                             ON CONFLICT (pk) DO UPDATE SET
                                 source_filename = EXCLUDED.source_filename,
                                 isa_name = EXCLUDED.isa_name,
                                 med_desc = EXCLUDED.med_desc,
                                 cycle_count_interval = EXCLUDED.cycle_count_interval,
                                 last_cycle_count = EXCLUDED.last_cycle_count,
                                 days_since_last_count = EXCLUDED.days_since_last_count,
                                 days_over_due = EXCLUDED.days_over_due;"""
                    execute_statement(sql, clean.to_dict("records"), batch=True, table_name="Cycle Count Status")

                elif u_type == "Cycle Count Variance Report":
                    clean = pd.concat([clean_cycle_count_variance_report(file) for file in uploaded_files], ignore_index=True)
                    sql = """INSERT INTO cycle_count_variances
                             (pk, variance_type, dt, med_id, med_desc, starting_qty, new_qty,
                              qty_variance, unit_cost, extended_cost, user_name, source_filename)
                             VALUES (%(pk)s, %(variance_type)s, %(dt)s, %(med_id)s, %(med_desc)s,
                                     %(starting_qty)s, %(new_qty)s, %(qty_variance)s, %(unit_cost)s,
                                     %(extended_cost)s, %(user_name)s, %(source_filename)s)
                             ON CONFLICT (pk) DO UPDATE SET
                                 variance_type = EXCLUDED.variance_type,
                                 dt = EXCLUDED.dt,
                                 med_id = EXCLUDED.med_id,
                                 med_desc = EXCLUDED.med_desc,
                                 starting_qty = EXCLUDED.starting_qty,
                                 new_qty = EXCLUDED.new_qty,
                                 qty_variance = EXCLUDED.qty_variance,
                                 unit_cost = EXCLUDED.unit_cost,
                                 extended_cost = EXCLUDED.extended_cost,
                                 user_name = EXCLUDED.user_name,
                                 source_filename = EXCLUDED.source_filename;"""
                    execute_statement(sql, clean.to_dict("records"), batch=True, table_name="Cycle Count Variances")

                elif u_type == "Buyer Formulary Listing Report":
                    clean = pd.concat([clean_buyer_formulary_listing_report(file) for file in uploaded_files], ignore_index=True)
                    sql = """INSERT INTO buyer_formulary_listing
                             (pk, snapshot_date, source_filename, location, med_id, med_desc, package_size,
                              min_qty, max_qty, qty, ndc, distributor, item_code, purchase_date, received_qty)
                             VALUES (%(pk)s, %(snapshot_date)s, %(source_filename)s, %(location)s, %(med_id)s,
                                     %(med_desc)s, %(package_size)s, %(min_qty)s, %(max_qty)s, %(qty)s,
                                     %(ndc)s, %(distributor)s, %(item_code)s, %(purchase_date)s, %(received_qty)s)
                             ON CONFLICT (pk) DO UPDATE SET
                                 source_filename = EXCLUDED.source_filename,
                                 location = EXCLUDED.location,
                                 med_desc = EXCLUDED.med_desc,
                                 package_size = EXCLUDED.package_size,
                                 min_qty = EXCLUDED.min_qty,
                                 max_qty = EXCLUDED.max_qty,
                                 qty = EXCLUDED.qty,
                                 ndc = EXCLUDED.ndc,
                                 distributor = EXCLUDED.distributor,
                                 item_code = EXCLUDED.item_code,
                                 purchase_date = EXCLUDED.purchase_date,
                                 received_qty = EXCLUDED.received_qty,
                                 uploaded_at = NOW();"""
                    execute_statement(sql, clean.to_dict("records"), batch=True, table_name="Buyer Formulary Listing")

                elif u_type == "Physical Inventory Report":
                    clean = pd.concat([clean_physical_inventory_report(file) for file in uploaded_files], ignore_index=True)
                    sql = """INSERT INTO physical_inventory_snapshots
                             (pk, snapshot_date, source_filename, isa_name, med_id, med_desc, min_qty,
                              max_qty, on_hand_qty, location, unit_cost, extended_cost)
                             VALUES (%(pk)s, %(snapshot_date)s, %(source_filename)s, %(isa_name)s, %(med_id)s,
                                     %(med_desc)s, %(min_qty)s, %(max_qty)s, %(on_hand_qty)s, %(location)s,
                                     %(unit_cost)s, %(extended_cost)s)
                             ON CONFLICT (pk) DO UPDATE SET
                                 source_filename = EXCLUDED.source_filename,
                                 isa_name = EXCLUDED.isa_name,
                                 med_desc = EXCLUDED.med_desc,
                                 min_qty = EXCLUDED.min_qty,
                                 max_qty = EXCLUDED.max_qty,
                                 on_hand_qty = EXCLUDED.on_hand_qty,
                                 unit_cost = EXCLUDED.unit_cost,
                                 extended_cost = EXCLUDED.extended_cost,
                                 uploaded_at = NOW();"""
                    execute_statement(sql, clean.to_dict("records"), batch=True, table_name="Physical Inventory")

                elif u_type == "Audit Transaction Detail RC":
                    if len(uploaded_files) > 1:
                        clean = pd.concat(
                            [
                                clean_audit_transaction_detail_rc(read_uploaded_tabular(file), file.name)
                                for file in uploaded_files
                            ],
                            ignore_index=True,
                        )
                    else:
                        clean = clean_audit_transaction_detail_rc(raw, uploaded_names)
                    sql = """INSERT INTO audit_transaction_detail_rc
                             (pk, care_area_name, location, station_name, source_system, dt, user_name,
                              user_id, user_type, priority_code, transaction_type, med_id, med_desc,
                              generic_name, med_class, therapeutic_class, drawer_subdrawer_pocket,
                              min_qty, max_qty, dispense_amount, qty, beginning_qty, ending_qty,
                              unit_cost, extended_cost, discrepancy, discrepancy_difference,
                              discrepancy_resolution_desc, discrepancy_reason, correction_quantity_before,
                              correction_quantity_after, correction, resolution_user, resolution_dt,
                              waste_amount, waste_reason, witness_user_name, override_reason, override_flag,
                              ordering_physician_present, attending_physician_present, source_filename)
                             VALUES
                             (%(pk)s, %(care_area_name)s, %(location)s, %(station_name)s, %(source_system)s,
                              %(dt)s, %(user_name)s, %(user_id)s, %(user_type)s, %(priority_code)s,
                              %(transaction_type)s, %(med_id)s, %(med_desc)s, %(generic_name)s,
                              %(med_class)s, %(therapeutic_class)s, %(drawer_subdrawer_pocket)s,
                              %(min_qty)s, %(max_qty)s, %(dispense_amount)s, %(qty)s, %(beginning_qty)s,
                              %(ending_qty)s, %(unit_cost)s, %(extended_cost)s, %(discrepancy)s,
                              %(discrepancy_difference)s, %(discrepancy_resolution_desc)s,
                              %(discrepancy_reason)s, %(correction_quantity_before)s,
                              %(correction_quantity_after)s, %(correction)s, %(resolution_user)s,
                              %(resolution_dt)s, %(waste_amount)s, %(waste_reason)s,
                              %(witness_user_name)s, %(override_reason)s, %(override_flag)s,
                              %(ordering_physician_present)s, %(attending_physician_present)s, %(source_filename)s)
                             ON CONFLICT (pk) DO UPDATE SET
                                 location = EXCLUDED.location,
                                 station_name = EXCLUDED.station_name,
                                 source_system = EXCLUDED.source_system,
                                 user_name = EXCLUDED.user_name,
                                 user_id = EXCLUDED.user_id,
                                 user_type = EXCLUDED.user_type,
                                 priority_code = EXCLUDED.priority_code,
                                 transaction_type = EXCLUDED.transaction_type,
                                 med_desc = EXCLUDED.med_desc,
                                 generic_name = EXCLUDED.generic_name,
                                 med_class = EXCLUDED.med_class,
                                 therapeutic_class = EXCLUDED.therapeutic_class,
                                 drawer_subdrawer_pocket = EXCLUDED.drawer_subdrawer_pocket,
                                 min_qty = EXCLUDED.min_qty,
                                 max_qty = EXCLUDED.max_qty,
                                 dispense_amount = EXCLUDED.dispense_amount,
                                 qty = EXCLUDED.qty,
                                 beginning_qty = EXCLUDED.beginning_qty,
                                 ending_qty = EXCLUDED.ending_qty,
                                 unit_cost = EXCLUDED.unit_cost,
                                 extended_cost = EXCLUDED.extended_cost,
                                 discrepancy = EXCLUDED.discrepancy,
                                 discrepancy_difference = EXCLUDED.discrepancy_difference,
                                 discrepancy_resolution_desc = EXCLUDED.discrepancy_resolution_desc,
                                 discrepancy_reason = EXCLUDED.discrepancy_reason,
                                 correction_quantity_before = EXCLUDED.correction_quantity_before,
                                 correction_quantity_after = EXCLUDED.correction_quantity_after,
                                 correction = EXCLUDED.correction,
                                 resolution_user = EXCLUDED.resolution_user,
                                 resolution_dt = EXCLUDED.resolution_dt,
                                 waste_amount = EXCLUDED.waste_amount,
                                 waste_reason = EXCLUDED.waste_reason,
                                 witness_user_name = EXCLUDED.witness_user_name,
                                 override_reason = EXCLUDED.override_reason,
                                 override_flag = EXCLUDED.override_flag,
                                 ordering_physician_present = EXCLUDED.ordering_physician_present,
                                 attending_physician_present = EXCLUDED.attending_physician_present,
                                 source_filename = EXCLUDED.source_filename,
                                 uploaded_at = NOW();"""
                    execute_statement(sql, clean.to_dict("records"), batch=True, table_name="Audit Transaction Detail RC")
                    cost_updates = build_med_cost_updates_from_audit_detail(clean)
                    if not cost_updates.empty:
                        sql_costs = """INSERT INTO med_costs (med_id, cost_per_unit)
                                       VALUES (%(med_id)s, %(cost_per_unit)s)
                                       ON CONFLICT (med_id) DO UPDATE SET
                                           cost_per_unit = EXCLUDED.cost_per_unit;"""
                        execute_statement(
                            sql_costs,
                            cost_updates.to_dict("records"),
                            batch=True,
                            table_name="Med Costs from Audit Detail",
                        )
                        st.info(f"Refreshed {len(cost_updates)} med cost price{'s' if len(cost_updates) != 1 else ''} from Audit Transaction Detail RC.")

                elif u_type == "Packaging Report":
                    clean = pd.concat([clean_packaging_report(file) for file in uploaded_files], ignore_index=True)
                    sql = """INSERT INTO packaged_meds
                             (pk, dispense_dt, reception_num, med_id, med_desc, dose_form, qty_per_pack, qoh,
                              manufacturer, ndc, mfg_lot_number, mfg_expire_date, device_id,
                              hospital_lot_number, hospital_expire_date, bud, packaged_by, confirmer)
                             VALUES (%(pk)s, %(dispense_dt)s, %(reception_num)s, %(med_id)s, %(med_desc)s,
                                     %(dose_form)s, %(qty_per_pack)s, %(qoh)s, %(manufacturer)s, %(ndc)s,
                                     %(mfg_lot_number)s, %(mfg_expire_date)s, %(device_id)s,
                                     %(hospital_lot_number)s, %(hospital_expire_date)s, %(bud)s,
                                     %(packaged_by)s, %(confirmer)s)
                             ON CONFLICT (pk) DO UPDATE SET
                                 med_desc = EXCLUDED.med_desc,
                                 dose_form = EXCLUDED.dose_form,
                                 qty_per_pack = EXCLUDED.qty_per_pack,
                                 qoh = EXCLUDED.qoh,
                                 manufacturer = EXCLUDED.manufacturer,
                                 ndc = EXCLUDED.ndc,
                                 mfg_lot_number = EXCLUDED.mfg_lot_number,
                                 mfg_expire_date = EXCLUDED.mfg_expire_date,
                                 device_id = EXCLUDED.device_id,
                                 hospital_lot_number = EXCLUDED.hospital_lot_number,
                                 hospital_expire_date = EXCLUDED.hospital_expire_date,
                                 bud = EXCLUDED.bud,
                                 packaged_by = EXCLUDED.packaged_by,
                                 confirmer = EXCLUDED.confirmer;"""
                    execute_statement(sql, clean.to_dict("records"), batch=True, table_name="Packaging Report")

                elif u_type in {"IV Room Workload", "IV Room Batching"}:
                    clean = clean_iv_room_report(raw)
                    sql = """INSERT INTO iv_room_workload
                             (pk, facility_name, order_lot_number, compound_type, num_preparations, dose_number,
                              drug_name, order_date, ordered_time, order_dt, completed_on, priority_name,
                              prepare_tat_minutes, prepared_by, approved_by, secondary_approved_by)
                             VALUES (%(pk)s, %(facility_name)s, %(order_lot_number)s, %(compound_type)s,
                                     %(num_preparations)s, %(dose_number)s, %(drug_name)s, %(order_date)s,
                                     %(ordered_time)s, %(order_dt)s, %(completed_on)s, %(priority_name)s,
                                     %(prepare_tat_minutes)s, %(prepared_by)s, %(approved_by)s,
                                     %(secondary_approved_by)s)
                             ON CONFLICT (pk) DO NOTHING;"""
                    execute_statement(sql, clean.to_dict("records"), batch=True, table_name=u_type)
                    processed_count = len(clean)

                elif u_type == "IV Room Workflow Detail":
                    clean = clean_iv_room_workflow_detail(raw, uploaded_names)
                    sql = """INSERT INTO iv_room_workflow_detail
                             (pk, facility_name, order_lot_number, dose_number, ordered_on,
                              prepared_by, approved_by, drug_name, workflow_name, workflow_step_type,
                              workflow_step_name, workflow_step_category, start_date, start_time, stop_time,
                              start_dt, stop_dt, total_duration_minutes, source_file)
                             VALUES (%(pk)s, %(facility_name)s, %(order_lot_number)s, %(dose_number)s,
                                     %(ordered_on)s, %(prepared_by)s, %(approved_by)s, %(drug_name)s,
                                     %(workflow_name)s, %(workflow_step_type)s, %(workflow_step_name)s,
                                     %(workflow_step_category)s, %(start_date)s, %(start_time)s, %(stop_time)s,
                                     %(start_dt)s, %(stop_dt)s, %(total_duration_minutes)s, %(source_file)s)
                             ON CONFLICT (pk) DO UPDATE SET
                                 facility_name = EXCLUDED.facility_name,
                                 order_lot_number = EXCLUDED.order_lot_number,
                                 dose_number = EXCLUDED.dose_number,
                                 ordered_on = EXCLUDED.ordered_on,
                                 prepared_by = EXCLUDED.prepared_by,
                                 approved_by = EXCLUDED.approved_by,
                                 drug_name = EXCLUDED.drug_name,
                                 workflow_name = EXCLUDED.workflow_name,
                                 workflow_step_type = EXCLUDED.workflow_step_type,
                                 workflow_step_name = EXCLUDED.workflow_step_name,
                                 workflow_step_category = EXCLUDED.workflow_step_category,
                                 start_date = EXCLUDED.start_date,
                                 start_time = EXCLUDED.start_time,
                                 stop_time = EXCLUDED.stop_time,
                                 start_dt = EXCLUDED.start_dt,
                                 stop_dt = EXCLUDED.stop_dt,
                                 total_duration_minutes = EXCLUDED.total_duration_minutes,
                                 source_file = EXCLUDED.source_file,
                                 uploaded_at = NOW();"""
                    execute_statement(sql, clean.to_dict("records"), batch=True, table_name=u_type)
                    processed_count = len(clean)

                elif u_type == "WCC Compounding Stats":
                    clean = clean_wcc_compounding_stats(raw, uploaded_names)
                    sql = """INSERT INTO wcc_compounding_stats
                             (pk, component_name, component_id, order_name, administration_dt, barcode_status, source_file)
                             VALUES (%(pk)s, %(component_name)s, %(component_id)s, %(order_name)s,
                                     %(administration_dt)s, %(barcode_status)s, %(source_file)s)
                             ON CONFLICT (pk) DO UPDATE SET
                                 component_name = EXCLUDED.component_name,
                                 component_id = EXCLUDED.component_id,
                                 order_name = EXCLUDED.order_name,
                                 administration_dt = EXCLUDED.administration_dt,
                                 barcode_status = EXCLUDED.barcode_status,
                                 source_file = EXCLUDED.source_file,
                                 uploaded_at = NOW();"""
                    execute_statement(sql, clean.to_dict("records"), batch=True, table_name="WCC Compounding Stats")
                    processed_count = len(clean)

                elif u_type in {"Cartfill Stats (All Areas)", "WCC Cartfill Stats"}:
                    clean = clean_wcc_cartfill_stats(raw, uploaded_names)
                    sql = """INSERT INTO wcc_cartfill_stats
                             (pk, report_start_date, report_end_date, order_medication, med_id,
                              ready_for_dispense_dt, admin_given_dt, prepared_dt, prep_or_dispense_user,
                              location, pharmacy, cartfill_area, source_file)
                             VALUES (%(pk)s, %(report_start_date)s, %(report_end_date)s, %(order_medication)s,
                                     %(med_id)s, %(ready_for_dispense_dt)s, %(admin_given_dt)s,
                                     %(prepared_dt)s, %(prep_or_dispense_user)s, %(location)s,
                                     %(pharmacy)s, %(cartfill_area)s, %(source_file)s)
                             ON CONFLICT (pk) DO UPDATE SET
                                 report_start_date = EXCLUDED.report_start_date,
                                 report_end_date = EXCLUDED.report_end_date,
                                 order_medication = EXCLUDED.order_medication,
                                 med_id = EXCLUDED.med_id,
                                 ready_for_dispense_dt = EXCLUDED.ready_for_dispense_dt,
                                 admin_given_dt = EXCLUDED.admin_given_dt,
                                 prepared_dt = EXCLUDED.prepared_dt,
                                 prep_or_dispense_user = EXCLUDED.prep_or_dispense_user,
                                 location = EXCLUDED.location,
                                 pharmacy = EXCLUDED.pharmacy,
                                 cartfill_area = EXCLUDED.cartfill_area,
                                 source_file = EXCLUDED.source_file,
                                 uploaded_at = NOW();"""
                    execute_statement(sql, clean.to_dict("records"), batch=True, table_name="Cartfill Stats")
                    processed_count = len(clean)

                elif u_type == "IV Overnight Cartfill Model":
                    workbooks = [clean_overnight_cartfill_workbook(file) for file in uploaded_files]
                    orders = pd.concat(
                        [workbook.get("orders", pd.DataFrame()) for workbook in workbooks],
                        ignore_index=True,
                    )
                    windows = pd.concat(
                        [workbook.get("windows", pd.DataFrame()) for workbook in workbooks],
                        ignore_index=True,
                    )
                    staffing = pd.concat(
                        [workbook.get("staffing", pd.DataFrame()) for workbook in workbooks],
                        ignore_index=True,
                    )

                    if not orders.empty:
                        sql_orders = """INSERT INTO overnight_iv_cartfill_orders
                                        (pk, order_id, order_medication, ready_for_dispense_dt, admin_given_dt,
                                         prepared_dt, prep_or_dispense_user, pharmacy, event_date,
                                         required_start_dt, prep_lead_hours, hold_hours, is_sjs_cleanroom)
                                        VALUES (%(pk)s, %(order_id)s, %(order_medication)s, %(ready_for_dispense_dt)s,
                                                %(admin_given_dt)s, %(prepared_dt)s, %(prep_or_dispense_user)s,
                                                %(pharmacy)s, %(event_date)s, %(required_start_dt)s,
                                                %(prep_lead_hours)s, %(hold_hours)s, %(is_sjs_cleanroom)s)
                                        ON CONFLICT (pk) DO NOTHING;"""
                        execute_statement(sql_orders, orders.to_dict("records"), batch=True, table_name="Overnight Cartfill Orders")
                        processed_count += len(orders)

                    if not windows.empty:
                        sql_windows = """INSERT INTO overnight_iv_cartfill_windows
                                         (pk, cartfill_name, time_processed_raw, doses_due, pharmacy)
                                         VALUES (%(pk)s, %(cartfill_name)s, %(time_processed_raw)s,
                                                 %(doses_due)s, %(pharmacy)s)
                                         ON CONFLICT (pk) DO NOTHING;"""
                        execute_statement(sql_windows, windows.to_dict("records"), batch=True, table_name="Overnight Cartfill Windows")
                        processed_count += len(windows)

                    if not staffing.empty:
                        sql_staffing = """INSERT INTO overnight_iv_staffing_model
                                          (pk, schedule_date, day_name, shift_name, weekend_shift_label,
                                           assigned_staff, is_weekend, is_placeholder)
                                          VALUES (%(pk)s, %(schedule_date)s, %(day_name)s, %(shift_name)s,
                                                  %(weekend_shift_label)s, %(assigned_staff)s, %(is_weekend)s,
                                                  %(is_placeholder)s)
                                          ON CONFLICT (pk) DO NOTHING;"""
                        execute_statement(sql_staffing, staffing.to_dict("records"), batch=True, table_name="Overnight Staffing Model")
                        processed_count += len(staffing)

                    clean = orders

                # 3. Success & Refresh
                if clean is not None:
                    clear_app_upload_caches()
                    elapsed = time.perf_counter() - upload_started
                    st.success(
                        f"Successfully uploaded {processed_count or len(clean)} records "
                        f"from {len(uploaded_files)} file{'s' if len(uploaded_files) != 1 else ''} "
                        f"in {elapsed:.1f} seconds."
                    )
                else:
                    st.warning("File type logic not yet implemented for this selection.")

            except Exception as e:
                st.error(f"Processing Error: {e}")
            finally:
                for file in folder_file_handles:
                    file.close()


    # --- EXECUTE DATA LOADER ---
    df_events = df_config = df_pharm = df_sched = df_att = pd.DataFrame()
    if "start_date" in locals() and "end_date" in locals():
        try:
            df_events, df_config, df_pharm, df_sched, df_att = load_data(start_date, end_date)
        except Exception as e:
            st.error(f"Failed to load data: {e}")

    # 1. OVERVIEW
    if selected_page == "📊 Overview":
        if demo_mode_enabled():
            render_demo_hub(start_date, end_date, df_events, df_pharm, df_sched, df_att)
            st.divider()
        if not df_events.empty:
            st.markdown("## 🏥 Executive Summary")
            st.caption(f"Date range: {start_date.strftime('%b %d, %Y')} → {end_date.strftime('%b %d, %Y')}")

            # ── Classify event types ──────────────────────────────────────────
            ev = df_events.copy()
            ev["_etype"] = ev["event_type"].astype(str).str.lower().str.strip()

            # Refills: Restock, Refill, Load (not unload/cancel)
            refills  = ev[
                ev["_etype"].str.contains(r"restock|refill|\bload\b|replenish", regex=True, na=False) &
                ~ev["_etype"].str.contains("cancel|unload|empty", na=False)
            ]
            # Outdates: Outdate, Expired, Override
            outdates = ev[
                ev["_etype"].str.contains(r"outdat|expir|override", regex=True, na=False) &
                ~ev["_etype"].str.contains("cancel", na=False)
            ]
            # Unloads: Unload, Empty return bin (not cancelled/eject)
            unloads  = ev[
                ev["_etype"].str.contains(r"unload|empty", regex=True, na=False) &
                ~ev["_etype"].str.contains("cancel|eject", na=False)
            ]
            # Everything except verify (for total tx count)
            real_tx  = ev[~ev["_etype"].str.contains("verify", na=False)]
            zero_verify_events = load_zero_verify_refill_gaps(start_date, end_date)
            zero_verify_summary = summarize_zero_verify_refill_gaps(zero_verify_events)
            avg_zero_refill_gap = zero_verify_events["hours_since_refill"].dropna().mean() if not zero_verify_events.empty else np.nan

            session_stats = ev.groupby("session_id").agg(total_time=("machine_time_sec", "sum"))
            avg_time      = session_stats["total_time"].mean()

            # ── Row 1 — Top-line KPIs ─────────────────────────────────────────
            m1, m2, m3, m4, m5, m6, m7 = st.columns(7)
            m1.metric("Total Transactions", f"{len(real_tx):,}")
            m2.metric("Refills / Restocks", f"{len(refills):,}")
            m3.metric("Unloads",            f"{len(unloads):,}")
            m4.metric("Outdates",           f"{len(outdates):,}")
            m5.metric("Active Techs",       df_events["user_name"].nunique())
            m6.metric("Avg Session",        seconds_to_mmss(avg_time))
            m7.metric("Discrepancies",      int(df_events["discrepancy_qty"].ne(0).sum()))

            if not zero_verify_events.empty:
                st.markdown("### Zero Verify Watch")
                z1, z2, z3, z4, z5 = st.columns(5)
                z1.metric("Zero Verify Events", f"{len(zero_verify_events):,}")
                z2.metric("Med/Device Pairs", f"{len(zero_verify_summary):,}")
                z3.metric("Meds Hit Zero", f"{zero_verify_events['med_id'].nunique():,}")
                z4.metric("Devices With Zero", f"{zero_verify_events['device'].nunique():,}")
                z5.metric(
                    "Avg Refill → Zero",
                    f"{avg_zero_refill_gap:.1f}h" if pd.notna(avg_zero_refill_gap) else "-",
                )

                with st.expander("Review meds verified as zero", expanded=False):
                    st.dataframe(
                        zero_verify_summary,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "device": st.column_config.TextColumn("Device"),
                            "med_id": st.column_config.TextColumn("Med ID"),
                            "med_desc": st.column_config.TextColumn("Medication"),
                            "zero_verifies": st.column_config.NumberColumn("Zero Verifies", format="%d"),
                            "first_seen": st.column_config.DatetimeColumn("First Seen", format="MM/DD/YY HH:mm"),
                            "last_seen": st.column_config.DatetimeColumn("Last Seen", format="MM/DD/YY HH:mm"),
                            "avg_hours_since_refill": st.column_config.NumberColumn("Avg Hours Since Refill", format="%.1f"),
                            "median_hours_since_refill": st.column_config.NumberColumn("Median Hours Since Refill", format="%.1f"),
                            "last_prior_refill": st.column_config.DatetimeColumn("Last Prior Refill", format="MM/DD/YY HH:mm"),
                            "verify_users": st.column_config.TextColumn("Verify Users"),
                            "prior_refill_users": st.column_config.TextColumn("Prior Refill Users"),
                        },
                    )
                with st.expander("Review each zero verify gap", expanded=False):
                    detail_cols = [
                        "dt", "user_name", "device", "med_id", "med_desc", "qty",
                        "prior_refill_dt", "prior_refill_user", "prior_refill_event_type",
                        "prior_refill_qty", "hours_since_refill",
                    ]
                    st.dataframe(
                        zero_verify_events[[col for col in detail_cols if col in zero_verify_events.columns]],
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "dt": st.column_config.DatetimeColumn("Zero Verify Time", format="MM/DD/YY HH:mm"),
                            "user_name": st.column_config.TextColumn("Verify User"),
                            "prior_refill_dt": st.column_config.DatetimeColumn("Prior Refill Time", format="MM/DD/YY HH:mm"),
                            "prior_refill_user": st.column_config.TextColumn("Prior Refill User"),
                            "prior_refill_qty": st.column_config.NumberColumn("Prior Refill Qty", format="%.0f"),
                            "hours_since_refill": st.column_config.NumberColumn("Hours Since Refill", format="%.1f"),
                        },
                    )
            else:
                st.caption("Zero Verify Watch: no Verify Inventory transactions with quantity 0 in this window.")

            st.divider()

            # ── Proactive Alerts Panel ────────────────────────────────────────
            with st.expander("🔔 Proactive Alerts", expanded=True):
                alerts = []

                if not zero_verify_events.empty:
                    alerts.append((
                        "Zero",
                        (
                            f"**{len(zero_verify_summary)} med/device pair(s) verified at quantity 0** "
                            f"across {zero_verify_events['device'].nunique()} device(s)."
                        ),
                        "warning",
                    ))

                # Stockout risk: meds with ending_qty == 0 in the window
                if "ending_qty" in df_events.columns and "med_desc" in df_events.columns:
                    last_inv = (
                        df_events.sort_values("dt")
                        .groupby(["device", "med_desc"])["ending_qty"]
                        .last()
                        .reset_index()
                    )
                    stockouts = last_inv[last_inv["ending_qty"] == 0]
                    if not stockouts.empty:
                        alerts.append(("🚨", f"**{len(stockouts)} medication(s) at zero inventory** — potential stockout.", "error"))

                # High discrepancy rate: > 2% of transactions
                disc_count = int(df_events["discrepancy_qty"].ne(0).sum())
                total_count = len(real_tx)
                if total_count > 0 and disc_count / total_count > 0.02:
                    pct = disc_count / total_count * 100
                    alerts.append(("⚠️", f"**Discrepancy rate is {pct:.1f}%** ({disc_count}/{total_count} transactions) — above 2% threshold.", "warning"))

                # Tardy alert: load from DB if sched/att data present
                if not df_sched.empty and not df_att.empty:
                    try:
                        _sched = df_sched.copy()
                        _att   = df_att.copy()
                        _sched["match_key"] = _sched["staff_name"].apply(normalize_name)
                        _att["match_key"]   = _att["raw_name"].apply(normalize_name)
                        _sched["date_obj"]  = pd.to_datetime(_sched["dt"]).dt.date
                        _att["date_obj"]    = pd.to_datetime(_att["dt_date"]).dt.date
                        _sched = _sched[~_sched["match_key"].isin(load_admin_users())]
                        _merged = pd.merge(_sched, _att, on=["match_key", "date_obj"], how="inner")
                        _merged["actual"]    = pd.to_datetime(_merged["start_dt"], errors="coerce")
                        _merged["scheduled"] = _merged.apply(lambda x: parse_shift_start(x["date_obj"], x["shift_type"]), axis=1)
                        _merged = _merged.dropna(subset=["actual", "scheduled"])
                        _merged["delay_min"] = (_merged["actual"] - _merged["scheduled"]).dt.total_seconds() / 60
                        tardy_count = int((_merged["delay_min"] > 5).sum())
                        repeat_offenders = (
                            _merged[_merged["delay_min"] > 5]
                            .groupby("staff_name").size()
                        )
                        repeat_count = int((repeat_offenders >= 3).sum())
                        if tardy_count > 0:
                            msg = f"**{tardy_count} tardy clock-in(s)** in this window."
                            if repeat_count > 0:
                                msg += f" {repeat_count} technician(s) with 3+ tardies."
                            alerts.append(("⏰", msg, "warning"))
                    except Exception:
                        pass

                if alerts:
                    for icon, msg, level in alerts:
                        if level == "error":
                            st.error(f"{icon} {msg}")
                        else:
                            st.warning(f"{icon} {msg}")
                else:
                    st.success("✅ No active alerts for this date range.")

            st.divider()

            # ── Daily Activity Bar Chart (Refills + Unloads + Outdates stacked) ──
            st.subheader("📅 Daily Transaction Activity")
            st.caption("Day-by-day breakdown of refills, unloads, and outdates across the selected range.")

            daily_rows = []
            for label, subset in [("Refill / Restock", refills), ("Unload", unloads), ("Outdate", outdates)]:
                if not subset.empty:
                    d = subset.copy()
                    d["_date"] = d["dt"].dt.date
                    counts = d.groupby("_date").size().reset_index(name="count")
                    counts["type"] = label
                    daily_rows.append(counts)

            if daily_rows:
                daily_df = pd.concat(daily_rows).sort_values("_date")
                fig_daily = px.bar(
                    daily_df,
                    x="_date", y="count",
                    color="type",
                    color_discrete_map={
                        "Refill / Restock": "#3b82f6",
                        "Unload":           "#8b5cf6",
                        "Outdate":          "#f97316",
                    },
                    barmode="group",
                    labels={"_date": "", "count": "Transactions", "type": ""},
                    text="count"
                )
                fig_daily.update_traces(textposition="outside", textfont_size=10)
                fig_daily.update_layout(
                    bargap=0.2,
                    bargroupgap=0.05,
                    height=400,
                    xaxis=dict(tickformat="%b %d", tickangle=-30),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0)
                )
                st.plotly_chart(fig_daily, use_container_width=True)
            else:
                # Only unload-type events in DB — show those as the daily chart
                ev["_date"] = ev["dt"].dt.date
                daily_all = (
                    ev[~ev["_etype"].str.contains("cancel|verify", na=False)]
                    .groupby(["_date", "event_type"])
                    .size()
                    .reset_index(name="count")
                    .sort_values("_date")
                )
                if not daily_all.empty:
                    fig_daily = px.bar(
                        daily_all,
                        x="_date", y="count",
                        color="event_type",
                        barmode="group",
                        labels={"_date": "", "count": "Transactions", "event_type": ""},
                        text="count"
                    )
                    fig_daily.update_traces(textposition="outside", textfont_size=10)
                    fig_daily.update_layout(
                        height=400,
                        xaxis=dict(tickformat="%b %d", tickangle=-30),
                        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0)
                    )
                    st.plotly_chart(fig_daily, use_container_width=True)
                    st.caption(
                        "💡 Only unload-type events found in this range. "
                        "Upload a Daily Transaction Report to see refills and outdates."
                    )

            st.divider()

            # ── Heavy Hitter Stats — 3 columns ───────────────────────────────
            st.subheader("🏋️ Heavy Hitter Stats")
            col_a, col_b, col_c = st.columns(3)

            # Column A — Top techs by refill volume (fallback to all tx if no refills)
            with col_a:
                st.markdown("**🔝 Top Techs by Refill Volume**")
                source = refills if not refills.empty else real_tx
                label  = "Refills" if not refills.empty else "Transactions"
                top_tech = (
                    source.groupby("user_name")
                    .size()
                    .reset_index(name=label)
                    .sort_values(label, ascending=False)
                    .head(10)
                )
                fig_tech = px.bar(
                    top_tech, x=label, y="user_name",
                    orientation="h",
                    labels={label: label, "user_name": ""},
                    color=label, color_continuous_scale="Blues",
                    text=label
                )
                fig_tech.update_traces(textposition="outside")
                fig_tech.update_layout(
                    yaxis={"categoryorder": "total ascending"},
                    coloraxis_showscale=False,
                    height=380, margin=dict(l=0, r=30, t=10, b=0)
                )
                st.plotly_chart(fig_tech, use_container_width=True)

            # Column B — Busiest devices
            with col_b:
                st.markdown("**🖥️ Busiest Devices**")
                top_dev = (
                    real_tx.groupby("device")
                    .size()
                    .reset_index(name="transactions")
                    .sort_values("transactions", ascending=False)
                    .head(10)
                )
                fig_dev = px.bar(
                    top_dev, x="transactions", y="device",
                    orientation="h",
                    labels={"transactions": "Transactions", "device": ""},
                    color="transactions", color_continuous_scale="Purples",
                    text="transactions"
                )
                fig_dev.update_traces(textposition="outside")
                fig_dev.update_layout(
                    yaxis={"categoryorder": "total ascending"},
                    coloraxis_showscale=False,
                    height=380, margin=dict(l=0, r=30, t=10, b=0)
                )
                st.plotly_chart(fig_dev, use_container_width=True)

            # Column C — Outdates & Unloads by tech (stacked)
            with col_c:
                st.markdown("**📦 Outdates & Unloads by Tech**")
                combined = []
                if not outdates.empty:
                    ot = outdates.groupby("user_name").size().reset_index(name="count")
                    ot["type"] = "Outdate"
                    combined.append(ot)
                if not unloads.empty:
                    ul = unloads.groupby("user_name").size().reset_index(name="count")
                    ul["type"] = "Unload"
                    combined.append(ul)

                if combined:
                    combo_df = pd.concat(combined)
                    # Keep only top 10 techs by total
                    top_techs = (
                        combo_df.groupby("user_name")["count"]
                        .sum()
                        .nlargest(10)
                        .index
                    )
                    combo_df = combo_df[combo_df["user_name"].isin(top_techs)]
                    fig_ou = px.bar(
                        combo_df,
                        x="count", y="user_name",
                        color="type",
                        orientation="h",
                        barmode="stack",
                        labels={"count": "Events", "user_name": "", "type": ""},
                        color_discrete_map={"Outdate": "#f97316", "Unload": "#8b5cf6"},
                        text="count"
                    )
                    fig_ou.update_traces(textposition="inside")
                    fig_ou.update_layout(
                        yaxis={"categoryorder": "total ascending"},
                        height=380, margin=dict(l=0, r=10, t=10, b=0),
                        legend=dict(orientation="h", yanchor="bottom", y=1.02)
                    )
                    st.plotly_chart(fig_ou, use_container_width=True)
                else:
                    st.info("No outdate or unload events in this range.")

        else:
            st.info("No Pyxis event data found for the selected date range. Upload a Daily Transaction Report to get started.")

    elif selected_page == "🛡️ Compliance":
        if not df_events.empty:
            disc_df = df_events[df_events['discrepancy_qty'] != 0].copy()
            zero_verify_events = load_zero_verify_refill_gaps(start_date, end_date)
            zero_verify_summary = summarize_zero_verify_refill_gaps(zero_verify_events)
            avg_zero_refill_gap = zero_verify_events["hours_since_refill"].dropna().mean() if not zero_verify_events.empty else np.nan
            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("Count Errors", len(disc_df))
            c3.metric("Zero Verify Events", len(zero_verify_events))
            c4.metric("Devices With Zero", zero_verify_events["device"].nunique() if not zero_verify_events.empty else 0)
            c5.metric("Avg Refill → Zero", f"{avg_zero_refill_gap:.1f}h" if pd.notna(avg_zero_refill_gap) else "-")
            if not disc_df.empty:
                disc_df['abs_variance'] = disc_df['discrepancy_qty'].abs() * disc_df['cost_per_unit']
                total_loss = disc_df['abs_variance'].sum()
                c2.metric("Variance Value (Risk)", f"${total_loss:,.2f}")
                st.dataframe(disc_df[['dt', 'user_name', 'device', 'med_desc', 'discrepancy_qty', 'discrepancy_reason', 'cost_per_unit', 'abs_variance']], use_container_width=True, column_config={"abs_variance": st.column_config.NumberColumn("Risk Value", format="$%.2f")})
            else:
                st.success("✅ Zero discrepancies found!")

            if not zero_verify_events.empty:
                st.subheader("Meds Verified As Zero")
                st.caption("Verify Inventory rows where the counted quantity was entered as 0.")
                st.dataframe(
                    zero_verify_summary,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "device": st.column_config.TextColumn("Device"),
                        "med_id": st.column_config.TextColumn("Med ID"),
                        "med_desc": st.column_config.TextColumn("Medication"),
                        "zero_verifies": st.column_config.NumberColumn("Zero Verifies", format="%d"),
                        "first_seen": st.column_config.DatetimeColumn("First Seen", format="MM/DD/YY HH:mm"),
                        "last_seen": st.column_config.DatetimeColumn("Last Seen", format="MM/DD/YY HH:mm"),
                        "avg_hours_since_refill": st.column_config.NumberColumn("Avg Hours Since Refill", format="%.1f"),
                        "median_hours_since_refill": st.column_config.NumberColumn("Median Hours Since Refill", format="%.1f"),
                        "last_prior_refill": st.column_config.DatetimeColumn("Last Prior Refill", format="MM/DD/YY HH:mm"),
                        "verify_users": st.column_config.TextColumn("Verify Users"),
                        "prior_refill_users": st.column_config.TextColumn("Prior Refill Users"),
                    },
                )

    # 6. LOAD/UNLOAD
    elif selected_page == "🚚 Load/Unload":
        if not df_events.empty:
            loads = df_events[df_events['event_type'].str.contains('load|unload', case=False, na=False)]
            st.dataframe(loads[['dt', 'user_name', 'device', 'event_type', 'med_desc', 'qty']], use_container_width=True)

